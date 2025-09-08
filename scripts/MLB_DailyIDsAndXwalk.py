#!/usr/bin/env python3
"""
MLB_WeeklyIdsAndXwalk.py — ultra-fast DK/FD site IDs + name crosswalk for MLB.

Speed tricks:
  • Workbook opened ONCE (read_only, data_only) and reused.
  • iter_rows(values_only=True) everywhere (no ws.cell calls).
  • Narrow column windows; projections limited to A:AZ.
  • Autodetect headers cheaply; falls back to letters if needed.
"""

from __future__ import annotations

import argparse, json, re, sys, shutil, tempfile, datetime, os
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from functools import lru_cache
import unicodedata as _u
from difflib import SequenceMatcher

# ------------------------- defaults/paths -------------------------

THIS = Path(__file__).resolve()
ROOT = THIS.parents[1]  # repo root (should contain /public)

DEFAULT_XLSM   = r"C:\Users\cpenn\Dropbox\Sports Models\MLB\MLB September 8th.xlsm"
DEFAULT_PROJ   = str(ROOT)
DEFAULT_CONFIG = str(ROOT / "scripts" / "configs" / "mlb_classic.json")

# ------------------------- small utils ---------------------------

def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

def _stage_copy_for_read(src: Path) -> Tuple[Path, Path]:
    """Copy workbook to temp so it can stay open in Excel while we read."""
    tmpdir = Path(tempfile.mkdtemp(prefix="mlb_weekly_"))
    dst = tmpdir / src.name
    shutil.copy2(src, dst)
    return dst, tmpdir

# ------------------------ header normalization --------------------

_HEADER_ALIASES = {
    "teamabbrev": "Team",
    "teamabbr": "Team",
    "dk sal": "DK Sal",
    "fd sal": "FD Sal",
    "dk pown %": "DK pOWN%",
    "fd pown %": "FD pOWN%",
    "pos": "Pos",
}
def _norm_header_label(s: str) -> str:
    t = (s or "").replace("\u00A0", " ").replace("\u202F", " ").strip()
    key = re.sub(r"\s+", " ", t).lower()
    return _HEADER_ALIASES.get(key, t)

def _dedup_headers(names) -> List[str]:
    seen, out = {}, []
    for i, raw in enumerate(list(names)):
        s = "" if raw is None else str(raw).strip()
        if s == "" or s.lower().startswith("unnamed"):
            s = f"col_{i+1}"
        key = s
        if key in seen:
            seen[key] += 1
            key = f"{key}__{seen[key]}"
        else:
            seen[key] = 0
        out.append(key)
    return out

def _excel_col_to_idx(label: str) -> int:
    s = re.sub(r"[^A-Za-z]", "", str(label)).upper()
    if not s: return 0
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1

def _slice_from_letters(min_letter: Optional[str], max_letter: Optional[str], ws: Worksheet) -> Tuple[int,int]:
    if not min_letter and not max_letter:
        return 1, ws.max_column
    min_c = 1 if not min_letter else _excel_col_to_idx(min_letter) + 1
    max_c = ws.max_column if not max_letter else _excel_col_to_idx(max_letter) + 1
    if min_c > max_c:
        min_c, max_c = max_c, min_c
    return min_c, max_c

# ------------------------ name normalization ----------------------

_SUFFIXES = {"jr","sr","ii","iii","iv","v"}

@lru_cache(maxsize=100_000)
def _strip_accents(s: str) -> str:
    return _u.normalize("NFKD", s).encode("ascii","ignore").decode("ascii")

@lru_cache(maxsize=100_000)
def _norm_name(s: str) -> str:
    s = _strip_accents(str(s or "")).lower()
    s = re.sub(r"[^\w\s-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

@lru_cache(maxsize=100_000)
def _base_key(s: str) -> str:
    parts = _norm_name(s).split()
    if parts and parts[-1].strip(".") in _SUFFIXES: parts = parts[:-1]
    return " ".join(parts)

@lru_cache(maxsize=100_000)
def _fi_last(s: str) -> str:
    parts = _base_key(s).split()
    return (parts[0][:1] + " " + parts[-1]).strip() if parts else ""

@lru_cache(maxsize=100_000)
def _last(s: str) -> str:
    parts = _base_key(s).split()
    return parts[-1] if parts else ""

# ------------------------ projections reader (FAST) ---------------

def read_literal_table_ws(ws: Worksheet,
                          header_row: int,
                          data_start_row: int,
                          limit_to_col: str = "AZ") -> pd.DataFrame:
    """Fast literal reader using iter_rows(values_only=True) within A:limit_to_col."""
    min_c, max_c = _slice_from_letters("A", limit_to_col, ws)
    # header
    hdr_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row,
                                 min_col=min_c, max_col=max_c, values_only=True))
    headers = _dedup_headers([_norm_header_label(v) if v is not None else "" for v in hdr_vals])

    # data
    out_rows = []
    blanks_in_a_row = 0
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row,
                            min_col=min_c, max_col=max_c, values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        if all(v == "" for v in vals):
            blanks_in_a_row += 1
            if blanks_in_a_row >= 3:
                break
            continue
        blanks_in_a_row = 0
        out_rows.append(vals)

    df = pd.DataFrame(out_rows, columns=headers)
    df = df.replace("", np.nan).dropna(how="all").fillna("")
    df = df.loc[:, ~(df.astype(str).eq("").all())]
    return df

def _read_proj_block(wb, sheet: str, header_row: int, data_start_row: int) -> pd.DataFrame:
    if sheet not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet]
    df = read_literal_table_ws(ws, header_row, data_start_row, limit_to_col="AZ")
    # normalize columns
    def _resolve(df, name):
        if name in df.columns: return name
        low = {c.lower(): c for c in df.columns}
        return low.get(name.lower())
    player_f = _resolve(df, "Player") or "Player"
    team_f   = _resolve(df, "Team")   or "Team"
    pos_f    = _resolve(df, "Pos") or _resolve(df, "Position") or "Pos"
    keep = df[[c for c in [player_f, team_f, pos_f] if c in df.columns]].copy()
    keep.columns = ["Player","Team","Pos"][:len(keep.columns)]
    return keep

# ------------------------ salary readers (AUTO & FAST) ------------

FD_NAME_CANDIDATES  = ["nickname", "name", "player", "player name"]
FD_FNAME, FD_LNAME  = "first name", "last name"
FD_ID_CANDIDATES    = ["id", "player id", "fd id"]
FD_TEAM_CANDIDATES  = ["team", "teamabbrev"]
FD_POS_CANDIDATES   = ["position", "pos"]

DK_NAME_CANDIDATES  = ["name", "player", "player name"]
DK_ID_CANDIDATES    = ["id", "player id", "dk id"]
DK_TEAM_CANDIDATES  = ["team", "teamabbrev"]
DK_POS_CANDIDATES   = ["position", "pos"]

def _find_header_row_fast(ws: Worksheet, max_scan: int = 10, max_col_letter: str = "AZ") -> Tuple[int, Dict[str,int], int, int]:
    """Scan top rows quickly to find header row and build header map."""
    min_c, max_c = _slice_from_letters("A", max_col_letter, ws)
    best_r, best_nonempty, best_hdr = 1, -1, {}
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = next(ws.iter_rows(min_row=r, max_row=r, min_col=min_c, max_col=max_c, values_only=True))
        nonempty = sum(1 for v in vals if v not in (None, ""))
        if nonempty > best_nonempty:
            best_nonempty = nonempty
            best_r = r
            hdr = {}
            for i, v in enumerate(vals, start=min_c):
                if v in (None, ""): continue
                label = _norm_header_label(str(v))
                hdr[label.lower()] = i  # 1-based index
            best_hdr = hdr
    return best_r, best_hdr, min_c, max_c

def _auto_salary_read_ws(ws: Worksheet, site: str, row_hard_cap: Optional[int]=None) -> List[Dict[str,str]]:
    """Autodetect using a single iter_rows pass."""
    header_row, hdr, min_c, max_c = _find_header_row_fast(ws, max_scan=10, max_col_letter="AZ")

    def col1(names: List[str]) -> Optional[int]:
        for k in names:
            if k in hdr: return hdr[k]
        return None

    if site.lower() == "fd":
        name_c = col1(FD_NAME_CANDIDATES)
        fname_c, lname_c = hdr.get(FD_FNAME), hdr.get(FD_LNAME)
        id_c   = col1(FD_ID_CANDIDATES)
        team_c = col1(FD_TEAM_CANDIDATES)
        pos_c  = col1(FD_POS_CANDIDATES)
    else:
        name_c = col1(DK_NAME_CANDIDATES)
        fname_c = lname_c = None
        id_c   = col1(DK_ID_CANDIDATES)
        team_c = col1(DK_TEAM_CANDIDATES)
        pos_c  = col1(DK_POS_CANDIDATES)

    if not id_c or (not name_c and not (fname_c and lname_c)):
        return []

    # Build a compact window just around the needed columns
    used = [c for c in [id_c, name_c, fname_c, lname_c, team_c, pos_c] if c]
    min_c2, max_c2 = min(used), max(used)

    out, seen_ids = [], set()
    BLANK_BREAK, blank_run, seen_any = 200, 0, False

    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row,
                            min_col=min_c2, max_col=max_c2, values_only=True):
        def get(abs_col: Optional[int]) -> str:
            if abs_col is None: return ""
            idx = abs_col - min_c2
            v = row[idx] if 0 <= idx < len(row) else None
            return "" if v in (None, "") else str(v).strip()

        pid = get(id_c)
        if not pid:
            if seen_any:
                blank_run += 1
                if blank_run >= BLANK_BREAK: break
            continue
        blank_run = 0; seen_any = True

        name = get(name_c) if name_c else (" ".join([get(fname_c), get(lname_c)]).strip())
        team = get(team_c).upper() if team_c else ""
        pos  = get(pos_c).upper()  if pos_c  else ""

        pid_clean = (re.search(r"([\d\-]+)", pid) or [pid])[0]
        if pid_clean in seen_ids:
            continue
        seen_ids.add(pid_clean)

        out.append({"name": name, "id": pid_clean, "team": team, "pos": pos})
        if row_hard_cap and len(out) >= int(row_hard_cap):
            break

    return out

def _col0(c: Union[str, int]) -> int:
    if isinstance(c, int): return max(0, c)
    s = re.sub(r"[^A-Za-z]", "", str(c)).upper()
    n = 0
    for ch in s: n = n*26 + (ord(ch)-64)
    return max(0, n - 1)

def _salary_read_sheet_letter_ws(ws: Worksheet, name_col, id_col, team_col=None, pos_col=None,
                                 row_hard_cap: Optional[int]=None) -> List[Dict[str,str]]:
    ni = _col0(name_col); ii = _col0(id_col)
    ti = _col0(team_col) if team_col is not None else None
    pi = _col0(pos_col)  if pos_col  is not None else None

    needed = [x for x in [ni, ii, ti, pi] if x is not None]
    min_c = min(needed) + 1
    max_c = max(needed) + 1

    out, seen_ids = [], set()
    blank_run, seen_any = 0, False
    BLANK_BREAK = 200

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=min_c, max_col=max_c, values_only=True):
        name = str(row[ni-(min_c-1)] or "").strip()
        pid  = str(row[ii-(min_c-1)] or "").strip()
        team = str(row[ti-(min_c-1)] or "").strip().upper() if ti is not None else ""
        pos  = str(row[pi-(min_c-1)] or "").strip().upper() if pi is not None else ""

        if (not name or name == "Name") and not pid:
            if seen_any:
                blank_run += 1
                if blank_run >= BLANK_BREAK:
                    break
            continue
        blank_run = 0
        seen_any = True
        if not name or not pid:
            continue

        pid_clean = (re.search(r"([\d\-]+)", pid) or [pid])[0]
        if pid_clean in seen_ids:
            continue
        seen_ids.add(pid_clean)

        out.append({"name": name, "id": pid_clean, "team": team, "pos": pos})
        if row_hard_cap and len(out) >= int(row_hard_cap):
            break
    return out

def _salary_read_sheet(wb, sheet: str, site: str, cfg_block: Dict[str,Any]) -> List[Dict[str,str]]:
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    row_cap = cfg_block.get("row_hard_cap")

    if site.lower() == "fd" and cfg_block.get("fd_autodetect", False):
        rows = _auto_salary_read_ws(ws, "fd", row_cap)
        if rows: return rows
    if site.lower() == "dk" and cfg_block.get("dk_autodetect", False):
        rows = _auto_salary_read_ws(ws, "dk", row_cap)
        if rows: return rows

    # fallback: letter mode (adjust letters if needed)
    if site.lower() == "fd":
        return _salary_read_sheet_letter_ws(
            ws,
            cfg_block.get("fd_name_col", "E"),
            cfg_block.get("fd_id_col",   "B"),
            cfg_block.get("fd_team_col"),
            cfg_block.get("fd_pos_col"),
            row_cap,
        )
    else:
        return _salary_read_sheet_letter_ws(
            ws,
            cfg_block.get("dk_name_col", "C"),
            cfg_block.get("dk_id_col",   "D"),
            cfg_block.get("dk_team_col"),
            cfg_block.get("dk_pos_col"),
            row_cap,
        )

# ------------------------ site IDs --------------------------------

def run_site_ids(wb, project_root: Path, cfg: Dict[str, Any]) -> None:
    scfg = cfg.get("site_ids")
    if not scfg:
        print("⚠️  site_ids config missing — skipping.")
        return

    out_rel = (scfg.get("out_rel") or "").lstrip(r"\\/")
    if not out_rel:
        print("⚠️  site_ids.out_rel missing — skipping."); return

    dk_rows = _salary_read_sheet(wb, scfg.get("dk_sheet", "DK Sals"), "dk", scfg)
    print(f"   DK site ids: {len(dk_rows)}")

    fd_rows = _salary_read_sheet(wb, scfg.get("fd_sheet", "FD Sals"), "fd", scfg)
    print(f"   FD site ids: {len(fd_rows)}")

    out = {"dk": dk_rows, "fd": fd_rows}
    out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
    ensure_parent(out_path)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✔️  JSON → {out_path}  (dk={len(dk_rows)}, fd={len(fd_rows)})")

# ------------------------ name crosswalk (FAST) -------------------

def run_name_xwalk(wb, project_root: Path, cfg: Dict[str, Any]) -> None:
    nx = cfg.get("name_xwalk")
    if not nx:
        print("⚠️  name_xwalk config missing — skipping.")
        return

    out_rel = (nx.get("out_rel") or "").lstrip(r"\\/")
    if not out_rel:
        print("⚠️  name_xwalk.out_rel missing — skipping."); return

    sheets = nx.get("projection_sheets") or [nx.get("projections_sheet","Batter Projections")]
    header_row = nx.get("header_row", 2)
    data_row   = nx.get("data_start_row", 3)

    proj_parts = []
    for sh in sheets:
        try:
            proj_parts.append(_read_proj_block(wb, sh, header_row, data_row))
        except Exception:
            pass

    if not proj_parts:
        print("⚠️  no projection sheets found to xwalk — aborting.")
        return

    proj = pd.concat(proj_parts, ignore_index=True).dropna(how="all").fillna("")
    if proj.empty:
        print("⚠️  empty projections after concat — aborting.")
        return

    # Precompute normalized keys (vectorized)
    proj["TeamU"] = proj["Team"].astype(str).str.upper()
    proj["PosU"]  = proj["Pos"].astype(str).str.upper().str.split("/").str[0].fillna("")
    proj["norm"]  = proj["Player"].map(_norm_name)
    proj["base"]  = proj["Player"].map(_base_key)
    proj["filast"]= proj["Player"].map(_fi_last)
    proj["last"]  = proj["Player"].map(_last)

    # Load salary rows (using autodetect where enabled)
    scfg = cfg.get("site_ids") or {}
    dk_rows = _salary_read_sheet(wb, scfg.get("dk_sheet","DK Sals"), "dk", scfg)
    fd_rows = _salary_read_sheet(wb, scfg.get("fd_sheet","FD Sals"), "fd", scfg)

    dk = pd.DataFrame(dk_rows); fd = pd.DataFrame(fd_rows)
    for df in (dk, fd):
        if df.empty:
            df["name"]=[]; df["id"]=[]; df["team"]=[]; df["pos"]=[]
        df["team"] = df["team"].astype(str).str.upper()
        df["pos"]  = df["pos"].astype(str).str.upper()
        df["norm"] = df["name"].map(_norm_name)
        df["base"] = df["name"].map(_base_key)
        df["filast"]=df["name"].map(_fi_last)
        df["last"] = df["name"].map(_last)

    # Build keyed maps → (key, team, pos) and key-only → rows
    def build_map_exact(df: pd.DataFrame, key: str):
        mp = {}
        for _, r in df.iterrows():
            k = (r[key], r["team"], r["pos"])
            if k not in mp:
                mp[k] = r
        return mp

    def build_map_keyonly(df: pd.DataFrame, key: str):
        mp = {}
        for _, r in df.iterrows():
            k = r[key]
            mp.setdefault(k, []).append(r)
        return mp

    dk_maps_exact = { "norm": build_map_exact(dk, "norm"),
                      "base": build_map_exact(dk, "base"),
                      "filast": build_map_exact(dk, "filast") }
    fd_maps_exact = { "norm": build_map_exact(fd, "norm"),
                      "base": build_map_exact(fd, "base"),
                      "filast": build_map_exact(fd, "filast") }
    dk_maps_key   = { "norm": build_map_keyonly(dk, "norm"),
                      "base": build_map_keyonly(dk, "base"),
                      "filast": build_map_keyonly(dk, "filast") }
    fd_maps_key   = { "norm": build_map_keyonly(fd, "norm"),
                      "base": build_map_keyonly(fd, "base"),
                      "filast": build_map_keyonly(fd, "filast") }

    dk_last_bucket = {ln: g.copy() for ln, g in dk.groupby("last")}
    fd_last_bucket = {ln: g.copy() for ln, g in fd.groupby("last")}

    def try_exact_maps(p_row, maps_exact, maps_keyonly) -> Optional[pd.Series]:
        team, pos = p_row["TeamU"], p_row["PosU"]
        for key in ("norm","base","filast"):
            hit = maps_exact[key].get((p_row[key], team, pos))
            if hit is not None: return hit
        for key in ("norm","base","filast"):
            hit = maps_exact[key].get((p_row[key], team, ""))
            if hit is not None: return hit
        for key in ("norm","base","filast"):
            hit = maps_exact[key].get((p_row[key], "", pos))
            if hit is not None: return hit
        for key in ("norm","base","filast"):
            rows = maps_keyonly[key].get(p_row[key], [])
            if rows: return rows[0]
        return None

    def fuzzy_from_bucket(p_row, bucket: dict, min_ratio=0.94) -> Optional[pd.Series]:
        if p_row["last"] not in bucket: return None
        cand = bucket[p_row["last"]]
        team = p_row["TeamU"]; pos = p_row["PosU"]
        if team:
            tmp = cand[cand["team"] == team]
            if not tmp.empty: cand = tmp
        if pos:
            tmp = cand[cand["pos"] == pos]
            if not tmp.empty: cand = tmp
        if cand.empty: return None
        target = _base_key(p_row["Player"])
        best, best_ratio = None, 0.0
        for _, rr in cand.iterrows():
            r = SequenceMatcher(None, target, rr["base"]).ratio()
            if r > best_ratio:
                best_ratio, best = r, rr
        return best if best_ratio >= min_ratio else None

    out_rows = []
    for _, p in proj.iterrows():
        pname = str(p.get("Player","")).strip()
        if not pname:
            continue
        dk_hit = try_exact_maps(p, dk_maps_exact, dk_maps_key)
        fd_hit = try_exact_maps(p, fd_maps_exact, fd_maps_key)
        if dk_hit is None: dk_hit = fuzzy_from_bucket(p, dk_last_bucket, 0.94)
        if fd_hit is None: fd_hit = fuzzy_from_bucket(p, fd_last_bucket, 0.94)
        out_rows.append({
            "proj": pname,
            "team": p["TeamU"],
            "pos":  p["PosU"],
            "dk_name": ("" if dk_hit is None else str(dk_hit.get("name",""))),
            "dk_id":   ("" if dk_hit is None else str(dk_hit.get("id",""))),
            "fd_name": ("" if fd_hit is None else str(fd_hit.get("name",""))),
            "fd_id":   ("" if fd_hit is None else str(fd_hit.get("id",""))),
        })

    out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
    ensure_parent(out_path)
    out_path.write_text(json.dumps(out_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✔️  JSON → {out_path}  (xwalk rows: {len(out_rows)})")

# ------------------------ projections merge (NEW) ------------------

def _read_json_file(path: Path):
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def _atomic_write_json(path: Path, payload):
    ensure_parent(path)
    fd, tmp = tempfile.mkstemp(dir=str(path.parent), prefix=".tmp_", suffix=".json")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        os.replace(tmp, path)
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

def run_merge_projections(project_root: Path, cfg: Dict[str, Any]) -> None:
    """
    Merge batters_projections.json + pitchers_projections.json -> projections.json
    Config (scripts/configs/mlb_classic.json):
      "projections_merge": {
        "batters_rel":  "data/mlb/latest/batters_projections",
        "pitchers_rel": "data/mlb/latest/pitchers_projections",
        "out_rel":      "data/mlb/latest/projections"
      }
    """
    pm = cfg.get("projections_merge")
    if not pm:
        print("ℹ️  projections_merge config missing — skipping merge.")
        return

    bat_rel = (pm.get("batters_rel") or "").lstrip(r"\\/")
    pit_rel = (pm.get("pitchers_rel") or "").lstrip(r"\\/")
    out_rel = (pm.get("out_rel")      or "").lstrip(r"\\/")
    if not (bat_rel and pit_rel and out_rel):
        print("⚠️  projections_merge.* missing paths — skipping.")
        return

    bat_p = (project_root / "public" / Path(bat_rel)).with_suffix(".json")
    pit_p = (project_root / "public" / Path(pit_rel)).with_suffix(".json")
    out_p = (project_root / "public" / Path(out_rel)).with_suffix(".json")

    if not bat_p.exists() or not pit_p.exists():
        print(f"⚠️  source projections not found (bat={bat_p.exists()} pit={pit_p.exists()}) — skipping.")
        return

    try:
        bat = _read_json_file(bat_p)
        pit = _read_json_file(pit_p)

        def rows(x):
            if isinstance(x, list): return x
            if isinstance(x, dict):
                for k in ("rows","players","data"):
                    v = x.get(k)
                    if isinstance(v, list):
                        return v
            return []

        merged = rows(bat) + rows(pit)
        _atomic_write_json(out_p, merged)
        print(f"✔️  projections merged → {out_p}  (bat={len(rows(bat))}, pit={len(rows(pit))}, total={len(merged)})")
    except Exception as e:
        print(f"ERROR merging projections: {e}", file=sys.stderr)

# ----------------------------- main --------------------------------

def _choose_project_root(arg_proj: Optional[str]) -> Path:
    if arg_proj:
        p = Path(arg_proj).resolve()
        if (p / "public").exists():
            return p
        print(f"⚠️  --project '{p}' has no /public. Falling back to script root.", file=sys.stderr)
    if (ROOT / "public").exists():
        return ROOT
    print("ERROR: could not find a project root with /public.", file=sys.stderr)
    sys.exit(1)

def main() -> None:
    ap = argparse.ArgumentParser(description="Build MLB site_ids and name_xwalk (ultra-fast).")
    ap.add_argument("--xlsm",    default=DEFAULT_XLSM,   help="Path to source workbook (.xls/.xlsx/.xlsm)")
    ap.add_argument("--project", default=DEFAULT_PROJ,   help="Path to project root (contains /public)")
    ap.add_argument("--config",  default=DEFAULT_CONFIG, help="Path to exporter config JSON")
    ap.add_argument("--only-site-ids", action="store_true", help="Only generate site_ids.json")
    ap.add_argument("--only-xwalk",    action="store_true", help="Only generate name_xwalk.json")
    args = ap.parse_args()

    xlsm_path    = Path(args.xlsm).resolve()
    project_root = _choose_project_root(args.project)
    config_path  = Path(args.config).resolve()

    if not xlsm_path.exists():
        print(f"ERROR: workbook not found: {xlsm_path}", file=sys.stderr); sys.exit(1)
    if not config_path.exists():
        print(f"ERROR: config not found: {config_path}", file=sys.stderr); sys.exit(1)

    # Stage copy so Excel can remain open; then open ONCE read-only
    staged_xlsm, temp_dir = _stage_copy_for_read(xlsm_path)
    try:
        cfg = json.loads(config_path.read_text(encoding="utf-8-sig"))
        wb = load_workbook(staged_xlsm, data_only=True, read_only=True, keep_links=False)

        if not args.only_xwalk:
            print("\n=== SITE IDS ===")
            run_site_ids(wb, project_root, cfg)

        if not args.only_site_ids:
            print("\n=== NAME XWALK ===")
            run_name_xwalk(wb, project_root, cfg)

        # NEW: always attempt to merge projections after the above
        print("\n=== MERGE PROJECTIONS ===")
        run_merge_projections(project_root, cfg)

        print("\nDone.")
        wb.close()
    finally:
        try: shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception: pass

if __name__ == "__main__":
    main()
