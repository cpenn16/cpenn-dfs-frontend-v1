#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NFL_Showdown_Exporter.py  (lean version, no Site IDs scraping)

- Exports workbook sheets per JSON config (tasks / cheatsheets / gameboard)
- (Optional) Merges existing DK/FD IDs + salaries + kickoff time from site_ids.json
  into projections.json (fast JSON-to-JSON pass; no Excel reading)

Run one-time IDs job:
  python NFL_Showdown_SiteIDs.py --xlsm "C:\\path\\NFL.xlsm" --project "." --config "scripts\\configs\\nfl_showdown.json"

Then use this fast exporter for frequent updates:
  python scripts/NFL_Showdown_Exporter.py --xlsm "C:\\path\\NFL.xlsm" --project "." --config "scripts\\configs\\nfl_showdown.json"

Skip merge if you don’t want to touch projections:
  ... --no-merge
"""

from __future__ import annotations

import argparse, json, re, sys, shutil, tempfile, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ---------- defaults ----------
THIS = Path(__file__).resolve()
ROOT = THIS.parents[1] if (len(THIS.parents) > 1) else THIS.parent
DEFAULT_XLSM   = r"C:\Users\cpenn\Dropbox\Sports Models\NFL\NFL SNF Showdown Bills vs Ravens.xlsm"
DEFAULT_PROJ   = str(ROOT)
DEFAULT_CONFIG = str(ROOT / "scripts" / "configs" / "nfl_showdown.json")

# ---------------- utilities ----------------
def ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)

def to_json_records(df: pd.DataFrame) -> str:
    df2 = df.astype(object).where(pd.notna(df), "")
    return df2.to_json(orient="records", force_ascii=False, indent=2)

def dedup(names: Iterable) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, raw in enumerate(list(names)):
        s = "" if raw is None else str(raw).strip()
        if s == "" or s.lower() in {"nan", "nat"} or s.lower().startswith("unnamed"):
            s = f"col_{i+1}"
        key = s
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        out.append(key)
    return out

def _stage_copy_for_read(src: Path) -> Tuple[Path, Path]:
    tmpdir = Path(tempfile.mkdtemp(prefix="nfl_showdown_"))
    dst = tmpdir / src.name
    shutil.copy2(src, dst)
    return dst, tmpdir

def _excel_col_to_idx(label: str) -> int:
    s = re.sub(r"[^A-Za-z]", "", str(label)).upper()
    if not s: return 0
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1

def _format_cell(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    fmt = cell.number_format or ""
    if isinstance(v, (datetime.date, datetime.datetime, datetime.time)):
        return str(v)
    if isinstance(v, (int, float, np.floating)):
        x = float(v)
        if "%" in str(fmt):
            n = x * 100.0 if abs(x) <= 1.01 else x
            return f"{n:.1f}%" if not float(n).is_integer() else f"{int(round(n))}%"
        return str(int(round(x))) if float(x).is_integer() else f"{x:.1f}"
    return str(v).strip()

def _norm_header_label(s: str) -> str:
    t = (s or "").replace("\u00A0", " ").replace("\u202F", " ").strip()
    key = re.sub(r"\s+", " ", t)
    return key

def _resolve_col(df: pd.DataFrame, name: str) -> Optional[str]:
    if name in df.columns:
        return name
    low_map = {c.lower(): c for c in df.columns}
    return low_map.get((name or "").lower())

# --------- light normalization (defensive: trims "%%" etc.) ----------
_PCT_LIKE = re.compile(r"%{2,}")
def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == object:
            out[c] = out[c].map(lambda v: _PCT_LIKE.sub("%", v) if isinstance(v, str) else v)
    return out

def read_literal_table(xlsm_path: Path, sheet: str,
                       header_row: Optional[int],
                       data_start_row: Optional[int],
                       limit_to_col: Optional[str] = None) -> pd.DataFrame:
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        ws = wb[sheet]

        max_c = ws.max_column
        if limit_to_col:
            max_c = min(max_c, _excel_col_to_idx(limit_to_col) + 1)

        if header_row is None or data_start_row is None:
            scan = min(8, ws.max_row)
            best_r, best_nonempty = 1, -1
            for r in range(1, scan + 1):
                vals = [c.value for c in ws[r][0:max_c]]
                nonempty = sum(1 for x in vals if x not in (None, ""))
                if nonempty > best_nonempty:
                    best_nonempty = nonempty
                    best_r = r
            header_row = best_r
            data_start_row = best_r + 1

        headers = dedup([_norm_header_label(_format_cell(c)) for c in ws[header_row][0:max_c]])

        out_rows: List[List[str]] = []
        blanks_in_a_row = 0
        for r in range(int(data_start_row), ws.max_row + 1):
            cells = ws[r][0:max_c]
            row = [_format_cell(c) for c in cells]
            if all(v == "" for v in row):
                blanks_in_a_row += 1
                if blanks_in_a_row >= 3: break
                continue
            blanks_in_a_row = 0
            out_rows.append(row)

        df = pd.DataFrame(out_rows, columns=headers)
        df = df.dropna(axis=0, how="all")
        df = df.replace("", np.nan).dropna(axis=0, how="all").fillna("")
        df = df.loc[:, ~(df.astype(str).eq("").all())]
        return df
    finally:
        wb.close()

# --------------- task runner ---------------
def maybe_apply_column_mapping(df: pd.DataFrame, mapping: Dict[str, str] | None) -> pd.DataFrame:
    if not mapping: return df
    existing = {src: dst for src, dst in mapping.items() if src in df.columns}
    return df.rename(columns=existing) if existing else df

def reorder_columns_if_all_present(df: pd.DataFrame, order: List[str] | None) -> pd.DataFrame:
    if not order: return df
    return df[order] if all(c in df.columns for c in order) else df

def _apply_leaf_filter(df: pd.DataFrame, f: Dict[str, Any]) -> pd.Series:
    name = _resolve_col(df, f.get("column", ""))
    if not name:
        return pd.Series([True] * len(df), index=df.index)
    op = (f.get("op") or "contains").lower()
    cs = bool(f.get("case_sensitive", False))
    s = df[name].astype(str)
    if not cs: s = s.str.lower()

    if op == "nonempty": return s.str.strip().ne("")
    val = str(f.get("value", "")).strip()
    if not cs: val = val.lower()

    if   op == "equals":       res = s.eq(val)
    elif op == "not_in":       res = ~s.isin([v.lower() if not cs else v for v in f.get("values", [])])
    elif op == "contains":     res = s.str.contains(val, na=False)
    elif op == "not_contains": res = ~s.str.contains(val, na=False)
    else:                      res = pd.Series([True] * len(df), index=df.index)
    return res.fillna(False)

def _apply_filters(df: pd.DataFrame, filters: Any) -> pd.DataFrame:
    if not filters: return df
    if isinstance(filters, list):
        masks = [_apply_leaf_filter(df, f) for f in filters]
        return df[pd.concat(masks, axis=1).all(axis=1)] if masks else df
    if isinstance(filters, dict):
        return df[_apply_leaf_filter(df, filters)]
    return df

def export_one(df: pd.DataFrame, out_csv: Optional[Path], out_json: Optional[Path]) -> None:
    if out_csv:
        ensure_parent(out_csv)
        df.astype(object).where(pd.notna(df), "").to_csv(out_csv, index=False, encoding="utf-8-sig")
        print(f"✔ CSV  → {out_csv}")
    if out_json:
        ensure_parent(out_json)
        out_json.write_text(to_json_records(df), encoding="utf-8")
        print(f"✔ JSON → {out_json}")

def run_task(xlsm_path: Path, project_root: Path, task: Dict[str, Any]) -> None:
    df = read_literal_table(
        xlsm_path=xlsm_path,
        sheet=task.get("sheet"),
        header_row=task.get("header_row"),
        data_start_row=task.get("data_start_row"),
        limit_to_col=task.get("limit_to_col"),
    )
    keep_cols = task.get("keep_columns_sheet_order") or []
    if keep_cols:
        df = df[[c for c in df.columns if c in keep_cols]]
    df = maybe_apply_column_mapping(df, task.get("column_mapping"))
    df = reorder_columns_if_all_present(df, task.get("column_order"))
    df = _apply_filters(df, task.get("filters"))

    # Light normalization for safety (removes "%%")
    df = normalize_df(df)

    out_rel = (task.get("out_rel") or "").lstrip(r"\/")
    fmt = str(task.get("format", "json")).lower()
    base = project_root / "public" / Path(out_rel)
    export_one(df,
               base.with_suffix(".csv") if fmt in ("csv", "both") else None,
               base.with_suffix(".json") if fmt in ("json", "both") else None)

# --------------- cheatsheets (optional) ---------------
def run_cheatsheets(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any]) -> None:
    cs = cfg.get("cheatsheets")
    if not cs: return
    sheet      = cs.get("sheet", "Cheat Sheet")
    out_rel    = (cs.get("out_rel") or "").lstrip(r"\/")
    title_ci   = bool(cs.get("title_match_ci", True))
    limit_rows = int(cs.get("limit_rows", 12))
    if not out_rel: return

    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            print("⚠ cheatsheets: sheet not found"); return
        ws = wb[sheet]

        def norm(s: Any) -> str:
            txt = "" if s is None else str(s).strip()
            return txt.lower() if title_ci else txt

        titles_cfg = cs.get("tables") or []
        all_titles_norm = {norm(t.get("title")) for t in titles_cfg if t.get("title")}
        index: Dict[str, List[Tuple[int,int]]] = {}
        for r in range(1, ws.max_row+1):
            for c in range(1, ws.max_column+1):
                s = norm(ws.cell(r,c).value)
                if s: index.setdefault(s, []).append((r,c))

        tables_out: List[Dict[str, Any]] = []
        for i, t in enumerate(titles_cfg):
            title = str(t.get("title") or f"Table {i+1}").strip()
            width = max(1, int(t.get("width", 3)))
            locs = index.get(norm(title), [])
            if not locs:
                print(f"⚠ cheatsheet title not found: {title}")
                continue
            start_r, start_c = min(locs, key=lambda rc: (rc[0], rc[1]))
            header_r = start_r
            data_r0  = header_r + 1
            hdr = [ws.cell(header_r, c) for c in range(start_c, min(start_c+width, ws.max_column+1))]
            headers = dedup([_norm_header_label(_format_cell(c)) for c in hdr])

            rows = []
            r = data_r0
            while r <= ws.max_row and len(rows) < limit_rows:
                row_cells = [ws.cell(r, c) for c in range(start_c, start_c+len(headers))]
                display = [_format_cell(c) for c in row_cells]
                if all(x == "" for x in display): break
                if norm(ws.cell(r, start_c).value) in all_titles_norm: break
                rows.append(display)
                r += 1

            sub = pd.DataFrame(rows, columns=headers)
            tables_out.append({
                "id": f"t{i+1}",
                "label": title,
                "columns": list(sub.columns),
                "rows": sub.astype(object).where(pd.notna(sub), "").to_dict(orient="records"),
            })

        out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
        ensure_parent(out_path)
        out_path.write_text(json.dumps({"tables": tables_out}, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✔ JSON → {out_path} (tables: {len(tables_out)})")
    finally:
        wb.close()

# --------------- gameboard (optional) ---------------
def run_gameboard(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any]) -> None:
    gb = cfg.get("gameboard")
    if not gb: return
    out_rel = (gb.get("out_rel") or "").lstrip(r"\/")
    if not out_rel: return
    title_re = re.compile(gb.get("title_regex", r"^\s*([A-Z]{2,4})\s*@\s*([A-Z]{2,4})\s*$"))
    yellow_rgbs = {str(x).upper() for x in gb.get("header_yellow_rgb", [])}

    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        sheet_name = None
        wants = gb.get("sheet")
        want_list = wants if isinstance(wants, list) else [wants]
        lower_map = {s.lower(): s for s in wb.sheetnames}
        for w in want_list:
            if not w: continue
            if w in wb.sheetnames: sheet_name = w; break
            if w.lower() in lower_map: sheet_name = lower_map[w.lower()]; break
        if not sheet_name:
            print("⚠ gameboard: sheet not found"); return

        ws = wb[sheet_name]

        def cell(r,c): 
            v = ws.cell(r,c).value
            return "" if v is None else str(v).strip()

        def is_yellow(r,c):
            try:
                f = ws.cell(r,c).fill
                if f and f.patternType == "solid":
                    rgb = (f.fgColor.rgb or "").upper()
                    return rgb in yellow_rgbs
            except Exception:
                pass
            return False

        games: List[Dict[str, Any]] = []
        for r in range(1, ws.max_row+1):
            # detect simple headers like "AAA @ BBB" (or colored)
            for c in range(1, ws.max_column+1):
                txt = cell(r,c)
                if not txt: continue
                if is_yellow(r,c) or title_re.match(txt):
                    m = title_re.match(txt)
                    if not m: continue
                    away, home = m.group(1), m.group(2)
                    # naive: grab next lines as blocks (kept minimal)
                    g = {"away": away, "home": home, "lines": []}
                    k = r+1
                    blanks=0
                    while k <= ws.max_row and len(g["lines"]) < 20:
                        rowtxt = " | ".join([cell(k, cc) for cc in range(c, min(c+12, ws.max_column+1)) if cell(k,cc)])
                        if not rowtxt:
                            blanks += 1
                            if blanks >= 2: break
                        else:
                            blanks = 0
                            g["lines"].append(rowtxt)
                        k += 1
                    games.append(g)
        out = (project_root / "public" / Path(out_rel)).with_suffix(".json")
        ensure_parent(out); out.write_text(json.dumps(games, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✔ JSON → {out} (games: {len(games)})")
    finally:
        wb.close()

# ---------- (Optional) merge site_ids.json into projections ----------
def _load_json(p: Path):
    try:
        return json.loads(p.read_text(encoding="utf-8-sig"))
    except FileNotFoundError:
        return None

def _write_json(p: Path, obj):
    ensure_parent(p)
    p.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def _to_rows_shape(raw):
    if raw is None:
        return [], ("none", None)
    if isinstance(raw, list):
        return raw, ("array", None)
    if isinstance(raw, dict):
        if isinstance(raw.get("rows"), list):
            return raw["rows"], ("rows", raw)
        if isinstance(raw.get("players"), list):
            return raw["players"], ("players", raw)
    return [], ("unknown", raw)

def _key(player: str, team: str) -> str:
    return f"{(player or '').strip().lower()}|{(team or '').strip().upper()}"

def _num(n: str | None) -> Optional[float]:
    if not n: return None
    s = str(n).replace("$","").replace(",","").strip()
    if s == "": return None
    try:
        return float(s)
    except Exception:
        return None

def _int_from_any(v) -> Optional[int]:
    if v is None: return None
    s = str(v).replace(",", "").replace("$", "").strip()
    try:
        return int(float(s))
    except Exception:
        return None

def merge_showdown_into_projections(project_root: Path, cfg: Dict[str, Any]) -> None:
    out_rel = "data/nfl/showdown/latest/projections"
    proj_path = (project_root / "public" / out_rel).with_suffix(".json")
    if not proj_path.exists():
        print(f"⚠ projections not found: {proj_path}"); return

    raw = _load_json(proj_path)
    rows, shape = _to_rows_shape(raw)

    # read site_ids.json (created by the standalone script)
    si_rel = (cfg.get("site_ids") or {}).get("out_rel") or "data/nfl/showdown/latest/site_ids"
    si_path = (project_root / "public" / Path(si_rel)).with_suffix(".json")
    si = _load_json(si_path) or {}
    dk_rows = si.get("dk", [])
    fd_rows = si.get("fd", [])
    dk_joined = si.get("dk_joined", {})  # preferred

    dk_flex_id  : Dict[str, str] = {}
    dk_cpt_id   : Dict[str, str] = {}
    dk_flex_sal : Dict[str, float] = {}
    dk_cpt_sal  : Dict[str, float] = {}
    kickoff_map : Dict[str, str] = {}

    if dk_joined:
        for k, v in dk_joined.items():
            kickoff_map.setdefault(k, v.get("time"))
            if v.get("flex", {}).get("id"):
                dk_flex_id[k] = v["flex"]["id"]
                dk_flex_sal[k] = _num(v["flex"].get("salary"))
            if v.get("cpt", {}).get("id"):
                dk_cpt_id[k] = v["cpt"]["id"]
                dk_cpt_sal[k] = _num(v["cpt"].get("salary"))
    else:
        for r in dk_rows:
            k = _key(r.get("name",""), r.get("team",""))
            kickoff_map.setdefault(k, r.get("time"))
            if (r.get("pos") or "").upper() == "CPT":
                dk_cpt_id[k]  = r.get("id","")
                dk_cpt_sal[k] = _num(r.get("salary"))
            else:
                dk_flex_id[k]  = r.get("id","")
                dk_flex_sal[k] = _num(r.get("salary"))

    fd_id_map   : Dict[str, str] = {}
    fd_flex_sal : Dict[str, float] = {}
    fd_mvp_sal  : Dict[str, float] = {}
    for r in fd_rows:
        k = _key(r.get("name",""), r.get("team",""))
        kickoff_map.setdefault(k, r.get("time"))
        fd_id_map[k]   = r.get("id","")
        fd_flex_sal[k] = _num(r.get("salary_flex"))
        fd_mvp_sal[k]  = _num(r.get("salary_mvp"))

    upd = dkf=dkc=fdf=t_hits=0
    for r in rows:
        player = r.get("player") or r.get("Player") or r.get("Player Name")
        team   = r.get("team")   or r.get("Team")   or r.get("TeamAbbrev")
        key    = _key(player, team)

        # IDs
        if key in dk_flex_id and dk_flex_id[key]:
            r["dk_flex_id"] = dk_flex_id[key]; dkf += 1
        if key in dk_cpt_id and dk_cpt_id[key]:
            r["dk_cpt_id"]  = dk_cpt_id[key];  dkc += 1
        if key in fd_id_map and fd_id_map[key]:
            r["fd_id"]      = fd_id_map[key];  fdf += 1

        # Salaries (printable strings)
        if key in dk_flex_sal and dk_flex_sal[key] is not None:
            r["DK Flex Sal"] = f"{int(dk_flex_sal[key]):,}"
        if key in dk_cpt_sal and dk_cpt_sal[key] is not None:
            r["DK CPT Sal"]  = f"{int(dk_cpt_sal[key]):,}"
        if key in fd_flex_sal and fd_flex_sal[key] is not None:
            r["FD Flex Sal"] = f"{int(fd_flex_sal[key]):,}"
        if key in fd_mvp_sal and fd_mvp_sal[key] is not None:
            r["FD MVP Sal"]  = f"{int(fd_mvp_sal[key]):,}"

        # kickoff time
        if not r.get("time") and key in kickoff_map and kickoff_map[key]:
            r["time"] = kickoff_map[key]; t_hits += 1

        # Fallback: compute CPT salary from DK Sal if still missing
        if not r.get("DK CPT Sal"):
            base = r.get("DK Sal") or r.get("dk_sal")
            base_num = _int_from_any(base)
            if base_num:
                r["DK CPT Sal"] = f"{int(base_num * 1.5):,}"

        # DK "MVP" aliases (display consistency)
        if r.get("dk_cpt_id") and not r.get("dk_mvp_id"):
            r["dk_mvp_id"] = r["dk_cpt_id"]
        if r.get("DK CPT Sal") and not r.get("DK MVP Sal"):
            r["DK MVP Sal"] = r["DK CPT Sal"]

        if (key in dk_flex_id) or (key in dk_cpt_id) or (key in fd_id_map):
            upd += 1

    _write_json(proj_path, rows if shape[0]=="array" else (shape[1] | {"rows": rows}))
    print("\n=== MERGE SHOWDOWN SALARIES/IDs INTO projections.json ===")
    print(f"• Projections updated: {upd}")
    print(f"• DK FLEX ids:        {dkf}")
    print(f"• DK CPT  ids:        {dkc}")
    print(f"• FD ids:             {fdf}")
    print(f"• Time hits:          {t_hits}")
    print(f"• Output:             {proj_path}")

# ---------------- main ----------------
def _choose_project_root(arg_proj: Optional[str]) -> Path:
    if arg_proj:
        p = Path(arg_proj).resolve()
        if (p / "public").exists(): return p
        print(f"⚠ --project '{p}' has no /public; using repo root.", file=sys.stderr)
    if (ROOT / "public").exists(): return ROOT
    print("ERROR: could not find a project root with /public.", file=sys.stderr)
    sys.exit(1)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsm",    default=DEFAULT_XLSM)
    ap.add_argument("--project", default=DEFAULT_PROJ)
    ap.add_argument("--config",  default=DEFAULT_CONFIG)
    ap.add_argument("--no-merge", action="store_true", help="Skip merging site_ids.json into projections.json")
    args = ap.parse_args()

    xlsm_path = Path(args.xlsm).resolve()
    project_root = _choose_project_root(args.project)
    cfg_path = Path(args.config).resolve()

    if not xlsm_path.exists():
        print(f"ERROR: workbook not found: {xlsm_path}", file=sys.stderr); sys.exit(1)
    if not cfg_path.exists():
        print(f"ERROR: config not found: {cfg_path}", file=sys.stderr); sys.exit(1)

    staged, tmpdir = _stage_copy_for_read(xlsm_path)
    try:
        cfg = json.loads(cfg_path.read_text(encoding="utf-8-sig"))
        # tasks
        for t in cfg.get("tasks", []):
            print(f"\n=== TASK: sheet='{t.get('sheet')}' | out='{t.get('out_rel')}' ===")
            try:
                run_task(staged, project_root, t)
            except Exception as e:
                print(f"⚠ task failed: {e}")

        # cheatsheets
        print("\n=== CHEAT SHEETS ===")
        try: run_cheatsheets(staged, project_root, cfg)
        except Exception as e: print(f"⚠ cheatsheets failed: {e}")

        # gameboard
        print("\n=== GAMEBOARD ===")
        try: run_gameboard(staged, project_root, cfg)
        except Exception as e: print(f"⚠ gameboard failed: {e}")

        # optional: merge IDs/salaries/time from JSON (no Excel; fast)
        if not args.no_merge:
            try:
                merge_showdown_into_projections(project_root, cfg)
            except Exception as e:
                print(f"⚠ merge failed: {e}")

        print("\nDone.")
    finally:
        try: shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception: pass

if __name__ == "__main__":
    main()
