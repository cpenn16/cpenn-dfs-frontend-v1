#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NFL_Showdown_SiteIDs.py — faster standalone exporter for DK/FD site IDs & salaries.

Speed-ups:
- iter_rows(values_only=True) over contiguous ranges (no Cell objects)
- Early stop after many blank names (--max-blank-rows)
- orjson dump if available; pretty-print only on request

Output JSON (same schema as before):
{
  "dk": [...], "fd": [...],
  "dk_joined": { "player|TEAM": { "name":..., "team":..., "flex":{id,salary}, "cpt":{id,salary}, "time": ... } }
}
"""

from __future__ import annotations

import argparse, re, sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---------- fast JSON ----------
try:
    import orjson
    def _dumps(obj, pretty: bool) -> bytes:
        opts = orjson.OPT_SERIALIZE_NUMPY
        if pretty: opts |= orjson.OPT_INDENT_2
        return orjson.dumps(obj, option=opts)
except Exception:  # pragma: no cover
    import json
    def _dumps(obj, pretty: bool) -> bytes:
        return (json.dumps(obj, ensure_ascii=False, indent=2 if pretty else None)).encode("utf-8")

from openpyxl import load_workbook

# ---------- defaults ----------
THIS = Path(__file__).resolve()
ROOT = THIS.parents[1] if (len(THIS.parents) > 1) else THIS.parent
DEFAULT_XLSM   = r"C:\Users\cpenn\Dropbox\Sports Models\NFL\NFL MNF Showdown Bears vs Vikings.xlsm"
DEFAULT_PROJ   = str(ROOT)
DEFAULT_CONFIG = str(ROOT / "scripts" / "configs" / "nfl_showdown.json")
DEFAULT_OUTREL = "data/nfl/showdown/latest/site_ids"

# ---------- helpers ----------
def ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)

def _excel_col_to_idx(label: str) -> int:
    """A→0, B→1, ...; ignores non-letters; returns 0-based index."""
    s = re.sub(r"[^A-Za-z]", "", str(label)).upper()
    if not s: return 0
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1

# DK role suffixes like "(CPT)", "- CPT", "(Flex)" → strip
_DK_ROLE_TOKENS_RE = re.compile(
    r"\s*(?:\((?:CPT|CAPT|Captain|FLEX)\)|-\s*(?:CPT|CAPT|Captain|FLEX)|\b(?:CPT|CAPT|Captain|FLEX)\b)\s*$",
    re.IGNORECASE,
)
def _norm_player_name_dk(name: str) -> str:
    s = (name or "").strip()
    prev = None
    while prev != s:
        prev = s
        s = _DK_ROLE_TOKENS_RE.sub("", s).strip()
    return re.sub(r"\s+", " ", s)

_TIME_RE = re.compile(r"\b(\d{1,2})\s*:\s*(\d{2})\s*([AP])\.?\s*M\b", re.I)
def _normalize_time(s: str | None) -> Optional[str]:
    if not s: return None
    m = _TIME_RE.search(str(s))
    if not m: return None
    h = int(m.group(1)); mm = m.group(2).zfill(2)
    ampm = "AM" if m.group(3).lower() == "a" else "PM"
    return f"{h}:{mm} {ampm}"

def _key(player: str, team: str) -> str:
    return f"{(player or '').strip().lower()}|{(team or '').strip().upper()}"

def _write_json(path: Path, obj, pretty: bool = True) -> None:
    ensure_parent(path)
    path.write_bytes(_dumps(obj, pretty))

# ---------- core ----------
def build_site_ids(
    xlsm_path: Path,
    project_root: Path,
    out_rel: str = DEFAULT_OUTREL,
    dk_sheet: str = "DK Salaries",
    fd_sheet: str = "FD Salaries",
    dk_name_col: str = "C",
    dk_id_col:   str = "D",
    dk_pos_col:  str = "E",
    dk_sal_col:  str = "F",
    dk_game_col: str = "G",
    dk_team_col: str = "H",
    fd_name_col: str = "D",
    fd_id_col:   str = "A",
    fd_pos_col:  str = "B",
    fd_sal_col:  str = "H",
    fd_mvp_col:  str = "I",
    fd_game_col: str = "J",
    fd_team_col: str = "K",
    max_blank_rows: int = 50,   # early stop threshold
) -> Path:
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)

    dk_rows: List[Dict[str, str]] = []
    fd_rows: List[Dict[str, str]] = []

    try:
        # -------- DraftKings (contiguous block C..H) --------
        if dk_sheet in wb.sheetnames:
            ws = wb[dk_sheet]
            cN = _excel_col_to_idx(dk_name_col)
            cI = _excel_col_to_idx(dk_id_col)
            cP = _excel_col_to_idx(dk_pos_col)
            cS = _excel_col_to_idx(dk_sal_col)
            cG = _excel_col_to_idx(dk_game_col)
            cT = _excel_col_to_idx(dk_team_col)

            start_c = min(cN, cI, cP, cS, cG, cT) + 1
            end_c   = max(cN, cI, cP, cS, cG, cT) + 1

            blanks = 0
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=start_c, max_col=end_c, values_only=True):
                name_raw = row[cN - (start_c-1)]
                if not name_raw:
                    blanks += 1
                    if blanks >= max_blank_rows: break
                    continue
                blanks = 0

                pid  = row[cI - (start_c-1)]
                if not pid:  # no ID → skip row
                    continue

                pos  = str(row[cP - (start_c-1)] or "").upper()
                sal  = str(row[cS - (start_c-1)] or "").strip()
                game = str(row[cG - (start_c-1)] or "").strip()
                team = str(row[cT - (start_c-1)] or "").upper()

                name = _norm_player_name_dk(str(name_raw))
                dk_rows.append({
                    "name": name,
                    "raw_name": str(name_raw),
                    "id": str(pid),
                    "team": team,
                    "pos": pos,            # FLEX or CPT
                    "salary": sal,
                    "game": game,
                    "time": _normalize_time(game)
                })

        # -------- FanDuel (contiguous block A..K; pick columns) --------
        if fd_sheet in wb.sheetnames:
            ws = wb[fd_sheet]
            c_use = [fd_id_col, fd_pos_col, fd_name_col, fd_sal_col, fd_mvp_col, fd_game_col, fd_team_col]
            idxs  = [_excel_col_to_idx(c) for c in c_use]
            start_c = min(idxs) + 1
            end_c   = max(idxs) + 1

            # mapping from absolute index to tuple position
            i_id   = _excel_col_to_idx(fd_id_col)   - (start_c-1)
            i_pos  = _excel_col_to_idx(fd_pos_col)  - (start_c-1)
            i_name = _excel_col_to_idx(fd_name_col) - (start_c-1)
            i_sal  = _excel_col_to_idx(fd_sal_col)  - (start_c-1)
            i_mvp  = _excel_col_to_idx(fd_mvp_col)  - (start_c-1)
            i_game = _excel_col_to_idx(fd_game_col) - (start_c-1)
            i_team = _excel_col_to_idx(fd_team_col) - (start_c-1)

            blanks = 0
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=start_c, max_col=end_c, values_only=True):
                name = row[i_name]
                if not name:
                    blanks += 1
                    if blanks >= max_blank_rows: break
                    continue
                blanks = 0

                pid  = row[i_id]
                if not pid:
                    continue

                pos  = str(row[i_pos] or "").upper()
                salF = str(row[i_sal] or "").strip()
                salM = str(row[i_mvp] or "").strip()
                game = str(row[i_game] or "").strip()
                team = str(row[i_team] or "").upper()

                fd_rows.append({
                    "name": str(name),
                    "id": str(pid),
                    "team": team,
                    "pos": pos,
                    "salary_flex": salF,
                    "salary_mvp":  salM,
                    "game": game,
                    "time": _normalize_time(game)
                })

    finally:
        wb.close()

    # -------- Build dk_joined (fast single pass) --------
    dk_joined: Dict[str, Dict[str, Any]] = {}
    for r in dk_rows:
        k = _key(r["name"], r["team"])
        d = dk_joined.get(k)
        if d is None:
            d = dk_joined[k] = {"name": r["name"], "team": r["team"], "flex": {}, "cpt": {}, "time": r.get("time")}
        if (r.get("pos") or "").upper() == "CPT":
            d["cpt"] = {"id": r["id"], "salary": r.get("salary")}
        else:
            d["flex"] = {"id": r["id"], "salary": r.get("salary")}
        if not d.get("time") and r.get("time"):
            d["time"] = r.get("time")

    # -------- Write JSON --------
    out_path = (Path(project_root) / "public" / Path(out_rel)).with_suffix(".json")
    payload = {"dk": dk_rows, "fd": fd_rows, "dk_joined": dk_joined}
    _write_json(out_path, payload, pretty=ARGS.pretty if 'ARGS' in globals() else False)
    print(f"✔ site_ids.json → {out_path}  (dk rows: {len(dk_rows)} | fd rows: {len(fd_rows)} | dk_joined: {len(dk_joined)})")
    return out_path

# ---------- CLI ----------
def parse_args():
    p = argparse.ArgumentParser(description="Build site_ids.json (DK/FD) for NFL Showdown from an XLSM workbook (fast).")
    p.add_argument("--xlsm",    default=DEFAULT_XLSM, help="Path to the Showdown XLSM workbook")
    p.add_argument("--project", default=DEFAULT_PROJ, help="Project root containing /public")
    p.add_argument("--config",  default=DEFAULT_CONFIG, help="Optional config JSON with site_ids settings")
    p.add_argument("--out-rel", default=None, help="Override output relative path (default: config -> site_ids.out_rel or fallback)")
    p.add_argument("--pretty",  action="store_true", help="Pretty-print JSON (slower & larger)")
    p.add_argument("--max-blank-rows", type=int, default=50, help="Early stop after this many blank-name rows")
    # Optional column/sheet overrides
    p.add_argument("--dk-sheet", default=None)
    p.add_argument("--fd-sheet", default=None)
    p.add_argument("--dk-name-col", default=None)
    p.add_argument("--dk-id-col",   default=None)
    p.add_argument("--dk-pos-col",  default=None)
    p.add_argument("--dk-sal-col",  default=None)
    p.add_argument("--dk-game-col", default=None)
    p.add_argument("--dk-team-col", default=None)
    p.add_argument("--fd-name-col", default=None)
    p.add_argument("--fd-id-col",   default=None)
    p.add_argument("--fd-pos-col",  default=None)
    p.add_argument("--fd-sal-col",  default=None)
    p.add_argument("--fd-mvp-col",  default=None)
    p.add_argument("--fd-game-col", default=None)
    p.add_argument("--fd-team-col", default=None)
    return p.parse_args()

def main():
    global ARGS
    ARGS = parse_args()
    xlsm_path = Path(ARGS.xlsm).resolve()
    project_root = Path(ARGS.project).resolve()
    if not xlsm_path.exists():
        print(f"ERROR: workbook not found: {xlsm_path}", file=sys.stderr); sys.exit(1)
    if not (project_root / "public").exists():
        print(f"ERROR: project root missing /public: {project_root}", file=sys.stderr); sys.exit(1)

    # Try to read optional config (only for defaults)
    si, out_rel = {}, None
    try:
        cfg = __import__("json").loads(Path(ARGS.config).read_text(encoding="utf-8-sig"))
        si = (cfg.get("site_ids") or {}) if isinstance(cfg, dict) else {}
        out_rel = si.get("out_rel")
    except Exception:
        pass

    out_rel = ARGS.out_rel or out_rel or DEFAULT_OUTREL
    dk_sheet = ARGS.dk_sheet or si.get("dk_sheet") or "DK Salaries"
    fd_sheet = ARGS.fd_sheet or si.get("fd_sheet") or "FD Salaries"

    build_site_ids(
        xlsm_path=xlsm_path,
        project_root=project_root,
        out_rel=out_rel,
        dk_sheet=dk_sheet,
        fd_sheet=fd_sheet,
        dk_name_col= ARGS.dk_name_col or si.get("dk_name_col", "C"),
        dk_id_col=   ARGS.dk_id_col   or si.get("dk_id_col",   "D"),
        dk_pos_col=  ARGS.dk_pos_col  or si.get("dk_pos_col",  "E"),
        dk_sal_col=  ARGS.dk_sal_col  or si.get("dk_sal_col",  "F"),
        dk_game_col= ARGS.dk_game_col or si.get("dk_game_col", "G"),
        dk_team_col= ARGS.dk_team_col or si.get("dk_team_col", "H"),
        fd_name_col= ARGS.fd_name_col or si.get("fd_name_col", "D"),
        fd_id_col=   ARGS.fd_id_col   or si.get("fd_id_col",   "A"),
        fd_pos_col=  ARGS.fd_pos_col  or si.get("fd_pos_col",  "B"),
        fd_sal_col=  ARGS.fd_sal_col  or si.get("fd_sal_col",  "H"),
        fd_mvp_col=  ARGS.fd_mvp_col  or si.get("fd_mvp_sal_col", "I"),
        fd_game_col= ARGS.fd_game_col or si.get("fd_game_col", "J"),
        fd_team_col= ARGS.fd_team_col or si.get("fd_team_col", "K"),
        max_blank_rows=ARGS.max_blank_rows,
    )

if __name__ == "__main__":
    main()
