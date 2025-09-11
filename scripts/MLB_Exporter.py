#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MLB_Exporter.py ‚Äî Full exporter to match NFL_Exporter structure (no _raw files)

Exports:
- Generic "tasks" (sheet ‚Üí out_rel) using literal Excel display values
- Cheat Sheet (table extraction by title cells, column-scoped so side-by-side tables don‚Äôt merge)
- MLB Matchups "gameboard" from the MLB Dashboard (panel windows like NFL)
- META: writes meta.json alongside any exported path (updated_iso/utc/epoch, source file & mtime)

Usage:
  python scripts/MLB_Exporter.py --xlsm "C:\\path\\to\\MLB.xlsm" ^
                                 --project "." ^
                                 --config "scripts\\configs\\mlb_classic.json"
"""

from __future__ import annotations

import argparse, json, re, sys, shutil, tempfile, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Union

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ------------------------- ROOT / DEFAULT PATHS -------------------------

THIS = Path(__file__).resolve()
ROOT = THIS.parents[1]  # repo root (expected to include /public)

DEFAULT_XLSM   = r"C:\Users\cpenn\Dropbox\Sports Models\MLB\MLB September 10th.xlsm"
DEFAULT_PROJ   = str(ROOT)
DEFAULT_CONFIG = str(ROOT / "scripts" / "configs" / "mlb_classic.json")


# ------------------------------ utilities ------------------------------

def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)

def dedup(names: Iterable) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, raw in enumerate(list(names)):
        s = "" if raw is None else str(raw).strip()
        if s == "" or s.lower() in {"nan", "nat"} or s.lower().startswith("unnamed"):
            s = f"col_{i+1}"
        key = s
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        out.append(key)
    return out

def to_json_records(df: pd.DataFrame) -> str:
    df2 = df.astype(object).where(pd.notna(df), "")
    return df2.to_json(orient="records", force_ascii=False, indent=2)

def _stage_copy_for_read(src: Path) -> tuple[Path, Path]:
    """Copy workbook to temp so Excel can stay open while we read."""
    tmpdir = Path(tempfile.mkdtemp(prefix="mlb_export_"))
    dst = tmpdir / src.name
    shutil.copy2(src, dst)
    return dst, tmpdir


# --------------------------- META (like NASCAR) ---------------------------

_META_DIRS: set[Path] = set()

def _mark_meta_dir(path_like: Optional[Path]) -> None:
    if not path_like:
        return
    p = Path(path_like).resolve()
    _META_DIRS.add(p.parent if p.is_file() else p)

def _write_meta_files(xlsm_path: Path) -> None:
    """Write meta.json into every directory in _META_DIRS."""
    if not _META_DIRS:
        return
    now_local = datetime.datetime.now().astimezone()
    now_utc   = now_local.astimezone(datetime.timezone.utc)
    src_name  = xlsm_path.name
    try:
        mtime = datetime.datetime.fromtimestamp(xlsm_path.stat().st_mtime).astimezone()
        mtime_iso = mtime.isoformat(timespec="seconds")
    except Exception:
        mtime_iso = None

    meta = {
        "updated_iso": now_local.isoformat(timespec="seconds"),
        "updated_utc": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "updated_epoch": int(time.time()),
        "source_file": src_name,
        "source_mtime_iso": mtime_iso,
    }

    for d in sorted(_META_DIRS):
        try:
            ensure_parent(d / "meta.json")
            (d / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"üïí meta  ‚Üí {d / 'meta.json'}")
        except Exception as e:
            print(f"‚ö†Ô∏è  meta write failed for {d}: {e}")


# --------------------- openpyxl ‚Äúdisplay text‚Äù formatting --------------------

_PERCENT_RE = re.compile(r"%")

def _decimals_from_format(fmt: str) -> int:
    if not isinstance(fmt, str):
        return 0
    m = re.search(r"0\.([0]+)", fmt)
    return len(m.group(1)) if m else 0

def _format_cell(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    fmt = cell.number_format or ""

    # Dates/times
    if isinstance(v, (datetime.date, datetime.datetime, datetime.time)):
        return str(v)

    # Numbers
    if isinstance(v, (int, float, np.floating)):
        x = float(v)
        if _PERCENT_RE.search(fmt):
            dec = _decimals_from_format(fmt)
            n = x * 100.0 if abs(x) <= 1.01 else x
            if float(n).is_integer():
                return f"{int(round(n))}%"
            return f"{n:.{dec}f}%"
        if float(x).is_integer():
            return str(int(round(x)))
        dec = _decimals_from_format(fmt) or 1
        return f"{x:.{dec}f}"

    return str(v).strip()


# ------------------------------ header normalization --------------------

_HEADER_ALIASES = {
    "dk sal": "DK Sal",
    "fd sal": "FD Sal",
    "teamabbrev": "Team",
    "matchup": "Opp",
    "opp": "Opp",
}

def _norm_header_label(s: str) -> str:
    t = (s or "").replace("\u00A0", " ").replace("\u202F", " ").strip()
    key = re.sub(r"\s+", " ", t).lower()
    return _HEADER_ALIASES.get(key, t)


# ------------------------------ filters engine --------------------------

def _resolve_col(df: pd.DataFrame, name: str) -> Optional[str]:
    if name in df.columns:
        return name
    low_map = {c.lower(): c for c in df.columns}
    return low_map.get((name or "").lower())

def _apply_leaf_filter(df: pd.DataFrame, f: Dict[str, Any]) -> pd.Series:
    col_name = _resolve_col(df, f.get("column", ""))
    if not col_name:
        return pd.Series([True] * len(df), index=df.index)

    op = (f.get("op") or "contains").lower()
    cs = bool(f.get("case_sensitive", False))
    s = df[col_name].astype(str)
    if not cs:
        s = s.str.lower()

    if op == "nonempty":       return s.str.strip().ne("")
    val = str(f.get("value", "")).strip()
    if not cs:                 val = val.lower()

    if   op == "equals":       res = s.eq(val)
    elif op == "not_equals":   res = s.ne(val)
    elif op == "contains":     res = s.str.contains(val, na=False)
    elif op == "not_contains": res = ~s.str.contains(val, na=False)
    elif op == "startswith":   res = s.str.startswith(val, na=False)
    elif op == "endswith":     res = s.str.endswith(val, na=False)
    elif op == "regex":
        try:
            pat = re.compile(val, 0 if cs else re.IGNORECASE)
            res = s.str.match(pat).fillna(False)
        except Exception:
            res = pd.Series([True] * len(df), index=df.index)
    else:
        res = pd.Series([True] * len(df), index=df.index)

    return res.fillna(False)

def _apply_filters(df: pd.DataFrame, filters: Union[List, Dict, None]) -> pd.DataFrame:
    def eval_filter(f) -> pd.Series:
        if isinstance(f, dict) and ("any_of" in f or "all_of" in f):
            if "any_of" in f:
                parts = [eval_filter(x) for x in (f.get("any_of") or [])]
                return pd.concat(parts, axis=1).any(axis=1) if parts else pd.Series([True]*len(df), index=df.index)
            if "all_of" in f:
                parts = [eval_filter(x) for x in (f.get("all_of") or [])]
                return pd.concat(parts, axis=1).all(axis=1) if parts else pd.Series([True]*len(df), index=df.index)
        return _apply_leaf_filter(df, f)

    if not filters: return df
    if isinstance(filters, dict) and ("any_of" in filters or "all_of" in filters):
        return df[eval_filter(filters)]
    if isinstance(filters, list):
        masks = [eval_filter(f) for f in filters]
        return df[pd.concat(masks, axis=1).all(axis=1)] if masks else df
    return df


# ------------------------------ task runner -----------------------------

def maybe_apply_column_mapping(df: pd.DataFrame, mapping: Dict[str, str] | None) -> pd.DataFrame:
    if not mapping: return df
    existing = {src: dst for src, dst in mapping.items() if src in df.columns}
    return df.rename(columns=existing) if existing else df

def reorder_columns_if_all_present(df: pd.DataFrame, order: List[str] | None) -> pd.DataFrame:
    if not order: return df
    return df[order] if all(c in df.columns for c in order) else df

def export_one(df: pd.DataFrame, out_csv: Optional[Path], out_json: Optional[Path]) -> None:
    if out_csv:
        ensure_parent(out_csv)
        df.astype(object).where(pd.notna(df), "").to_csv(out_csv, index=False, encoding="utf-8-sig")
        print(f"‚úîÔ∏è  CSV  ‚Üí {out_csv}")
        _mark_meta_dir(out_csv)
    if out_json:
        ensure_parent(out_json)
        out_json.write_text(to_json_records(df), encoding="utf-8")
        print(f"‚úîÔ∏è  JSON ‚Üí {out_json}")
        _mark_meta_dir(out_json)

def run_task(xlsm_path: Path, project_root: Path, task: Dict[str, Any]) -> None:
    sheet = task.get("sheet")
    if not sheet:
        print("‚ö†Ô∏è  SKIP: task missing 'sheet'"); return

    df = read_literal_table(
        xlsm_path=xlsm_path,
        sheet=sheet,
        header_row=task.get("header_row"),
        data_start_row=task.get("data_start_row"),
        limit_to_col=task.get("limit_to_col"),
    )

    keep_cols_src: List[str] = task.get("keep_columns_sheet_order", [])
    if keep_cols_src:
        df = df[[c for c in df.columns if c in keep_cols_src]]

    df = maybe_apply_column_mapping(df, task.get("column_mapping"))
    df = reorder_columns_if_all_present(df, task.get("column_order"))
    df = _apply_filters(df, task.get("filters"))

    out_rel = (task.get("out_rel") or "").lstrip(r"\/")
    if not out_rel:
        print(f"‚ö†Ô∏è  SKIP: task for '{sheet}' missing 'out_rel'"); return

    base = project_root / "public" / Path(out_rel)
    fmt = str(task.get("format", "json")).lower()
    csv_path  = base.with_suffix(".csv")  if fmt in ("csv", "both")  else None
    json_path = base.with_suffix(".json") if fmt in ("json", "both") else None
    export_one(df, csv_path, json_path)


# ------------------------------- literal read ---------------------------

def _excel_col_to_idx(label: str) -> int:
    s = re.sub(r"[^A-Za-z]", "", str(label)).upper()
    if not s: return 0
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1

def read_literal_table(xlsm_path: Path, sheet: str,
                       header_row: Optional[int],
                       data_start_row: Optional[int],
                       limit_to_col: Optional[str] = None) -> pd.DataFrame:
    """
    Read a sheet using openpyxl and return a DataFrame of *strings* matching Excel display.
    `limit_to_col` (e.g., "AE") caps the rightmost column read.
    """
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        ws = wb[sheet]

        max_c = ws.max_column
        if limit_to_col:
            try:
                max_c = min(max_c, _excel_col_to_idx(limit_to_col) + 1)
            except Exception:
                pass

        if header_row is None or data_start_row is None:
            scan = min(8, ws.max_row)
            best_r, best_nonempty = 1, -1
            for r in range(1, scan + 1):
                vals = [c.value for c in ws[r][0:max_c]]
                nonempty = sum(1 for x in vals if x not in (None, ""))
                if nonempty > best_nonempty:
                    best_nonempty = nonempty
                    best_r = r
            header_row = best_r
            data_start_row = best_r + 1

        raw_headers = [_format_cell(c) for c in ws[header_row][0:max_c]]
        raw_headers = [_norm_header_label(h) for h in raw_headers]
        headers = dedup(raw_headers)

        out_rows: List[List[str]] = []
        blanks_in_a_row = 0
        for r in range(int(data_start_row), ws.max_row + 1):
            cells = ws[r][0:max_c]
            row = [_format_cell(c) for c in cells]
            if all(v == "" for v in row):
                blanks_in_a_row += 1
                if blanks_in_a_row >= 2: break
                continue
            blanks_in_a_row = 0
            out_rows.append(row)

        df = pd.DataFrame(out_rows, columns=headers)
        df = df.dropna(axis=0, how="all")
        df = df.replace("", np.nan).dropna(axis=0, how="all").fillna("")
        df = df.loc[:, ~(df.astype(str).eq("").all())]
        return df
    finally:
        wb.close()


# -------------------------- cheatsheets (by title) ----------------------

def run_cheatsheets(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any]) -> None:
    """
    Title-based extraction, column-scoped.
    - Auto-detect header row (either the yellow title row or the row below).
    - If 'Player' is blank (HYPERLINK formula), pull display text from formula.
    - Normalize 'Time' to 12-hour (H:MM AM/PM).
    """
    cs = cfg.get("cheatsheets")
    if not cs:
        return
    sheet      = cs.get("sheet") or "Cheat Sheet"
    out_rel    = (cs.get("out_rel") or "").lstrip(r"\/")
    title_ci   = bool(cs.get("title_match_ci", True))
    default_limit = int(cs.get("limit_rows", 200))
    if not out_rel:
        print("‚ö†Ô∏è  SKIP cheatsheets: missing out_rel")
        return

    # helpers --------------------------------------------------------------
    _TIME12_RE = re.compile(r"^\s*(\d{1,2})\s*:\s*(\d{2})(?::\d{2})?\s*(AM|PM)?\s*$", re.I)
    def _to_12h(s: Any) -> str:
        if s is None: return ""
        t = str(s).strip()
        if not t: return ""
        m = _TIME12_RE.match(t)
        if not m: return t
        hh = int(m.group(1)); mm = m.group(2); ampm = m.group(3)
        if ampm: return f"{hh}:{mm} {ampm.upper()}"
        if hh == 0:  return f"12:{mm} AM"
        if hh <= 11: return f"{hh}:{mm} AM"
        if hh == 12:return f"12:{mm} PM"
        return f"{hh-12}:{mm} PM"

    _HL_RE = re.compile(r"^=\s*HYPERLINK\s*\(\s*(?:\"[^\"]*\"|[^,]+)\s*,\s*\"([^\"]+)\"\s*\)\s*$", re.I)
    def _hyperlink_display(val: Any) -> Optional[str]:
        s = "" if val is None else str(val)
        m = _HL_RE.match(s)
        return m.group(1).strip() if m else None

    def norm(s: Any) -> str:
        txt = "" if s is None else str(s).strip()
        return txt.lower() if title_ci else txt

    # ---------------------------------------------------------------------
    wb_data = load_workbook(xlsm_path, data_only=True,  read_only=True, keep_links=False)
    wb_form = load_workbook(xlsm_path, data_only=False, read_only=True, keep_links=False)
    try:
        if sheet not in wb_data.sheetnames:
            print(f"‚ö†Ô∏è  SKIP cheatsheets: sheet '{sheet}' not found")
            return
        ws  = wb_data[sheet]
        wsf = wb_form[sheet]
        n_rows, n_cols = ws.max_row, ws.max_column

        titles_cfg = cs.get("tables") or []
        all_titles_norm = {norm(t.get("title")) for t in titles_cfg if t.get("title")}

        # index of text ‚Üí [(r,c)]
        index: Dict[str, List[tuple]] = {}
        for r in range(1, n_rows + 1):
            for c in range(1, n_cols + 1):
                s = norm(ws.cell(r, c).value)
                if s:
                    index.setdefault(s, []).append((r, c))

        EXPECTED = {"player","salary","team","matchup","vegas","time","proj","value","pown"}

        def pick_header_row(start_r: int, start_c: int, width: int) -> int:
            """Choose between start_r or start_r+1 by scoring header-likeness."""
            candidates = [start_r, start_r + 1]
            best_r, best_score = start_r, -1
            for r0 in candidates:
                cells = [ws.cell(r0, c) for c in range(start_c, min(start_c+width, n_cols+1))]
                labels = [ _norm_header_label(_format_cell(c)) for c in cells ]
                labels_l = [l.lower() for l in labels]
                score = sum(1 for l in labels_l if l in EXPECTED)
                if "player" in labels_l: score += 3
                score += sum(1 for l in labels_l if l.strip() != "")
                if score > best_score:
                    best_r, best_score = r0, score
            return best_r

        out_obj: Dict[str, Any] = {}

        for i, t in enumerate(titles_cfg):
            title = str(t.get("title") or f"Table {i+1}").strip()
            width = max(1, int(t.get("width", 8)))
            limit_rows = int(t.get("limit_rows", default_limit))

            # all occurrences of this title
            locs = sorted(index.get(norm(title), []), key=lambda rc: (rc[0], rc[1]))
            if not locs:
                print(f"‚ö†Ô∏è  cheatsheets: title not found: '{title}'")
                continue

            for (start_r, start_c) in locs:
                header_r = pick_header_row(start_r, start_c, width)
                data_r0  = header_r + 1

                # headers within span
                hdr = [ws.cell(header_r, c) for c in range(start_c, min(start_c + width, n_cols + 1))]
                headers = dedup([_norm_header_label(_format_cell(c)) for c in hdr])

                # locate special columns
                col_l = [h.lower() for h in headers]
                idx_player = col_l.index("player") if "player" in col_l else None
                idx_time   = col_l.index("time")   if "time"   in col_l else None

                # -------- rows in this occurrence --------
                rows = []
                r = data_r0
                blanks = 0
                while r <= n_rows and len(rows) < limit_rows:
                    # stop when a new section title appears in the first cell (any title)
                    first = norm(ws.cell(r, start_c).value)
                    if first and first in all_titles_norm:
                        break

                    row_cells = [ws.cell(r, c) for c in range(start_c, start_c + len(headers))]
                    display   = [_format_cell(c) for c in row_cells]

                    # fill 'Player' from formula if needed
                    if idx_player is not None and not display[idx_player]:
                        raw = wsf.cell(r, start_c + idx_player).value
                        disp = _hyperlink_display(raw)
                        if disp:
                            display[idx_player] = disp

                    # normalize time
                    if idx_time is not None and display[idx_time]:
                        display[idx_time] = _to_12h(display[idx_time])

                    if all(x == "" for x in display):
                        blanks += 1
                        if blanks >= 2:
                            break
                        r += 1
                        continue

                    blanks = 0
                    rows.append(display)
                    r += 1

                # Build frame for this occurrence
                sub = pd.DataFrame(rows, columns=headers)

                # ==== normalization (unchanged) ====
                _PLAYER_SECTIONS = {"Pitcher","C","1B","2B","3B","SS","OF","Cash Core"}
                cols = list(sub.columns)
                if title in _PLAYER_SECTIONS and cols and cols[0] == title:
                    cols[0] = "Player"
                cols = ["Opp" if (c or "").lower() in ("matchup","opp") else c for c in cols]
                if title == "Top Stacks":
                    if cols and cols[0] in ("Top Stacks","Stacks","Stack"):
                        cols[0] = "Team"
                    seen_first_team = False
                    for i2 in range(1, len(cols)):
                        if not seen_first_team and cols[i2] == "Team":
                            cols[i2] = "Opp"
                            seen_first_team = True
                            continue
                        if seen_first_team and cols[i2] == "Opp":
                            cols[i2] = "Opp Pitcher"
                            break
                sub.columns = cols
                # ===================================

                recs = sub.astype(object).where(pd.notna(sub), "").to_dict(orient="records")

                # merge multiple occurrences (e.g., all three "OF" blocks)
                if title in out_obj and isinstance(out_obj[title], list):
                    out_obj[title].extend(recs)
                else:
                    out_obj[title] = recs

        out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
        ensure_parent(out_path)
        out_path.write_text(json.dumps(out_obj, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"‚úîÔ∏è  JSON ‚Üí {out_path}  (sections: {', '.join(out_obj.keys()) or 'none'})")
        _mark_meta_dir(out_path)
    finally:
        wb_data.close()
        wb_form.close()


# ---------------------- MLB GAMEBOARD (Dashboard) ‚Äî FAST ----------------------

# Flexible parser: pull AAA and BBB anywhere in the string (not anchored)
_HEADER_PAT = re.compile(r"([A-Z]{2,4})\s*@\s*([A-Z]{2,4})")

def _parse_header(text: str) -> tuple[str, str] | None:
    if not text:
        return None
    m = _HEADER_PAT.search(text.strip().upper())
    if not m:
        return None
    a, b = m.group(1), m.group(2)
    if 2 <= len(a) <= 4 and 2 <= len(b) <= 4:
        return a, b
    return None

def _build_grid(ws: Worksheet, max_rows: int, max_cols: int) -> list[list[str]]:
    """One pass streaming read ‚Üí in-memory grid of strings."""
    grid = []
    ncols = min(max_cols, ws.max_column or 1)
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row or 1),
                            min_col=1, max_col=ncols, values_only=True):
        grid.append([("" if v is None else str(v).strip()) for v in row])
    return grid

def _row_has_any_text(grid: list[list[str]], r: int, c0: int = 0, c1: int | None = None) -> bool:
    row = grid[r]
    if c1 is None: c1 = len(row) - 1
    c1 = min(c1, len(row) - 1)
    for c in range(max(0, c0), c1 + 1):
        if row[c]:
            return True
    return False

def _row_text_slice(grid: list[list[str]], r: int, c0: int, c1: int) -> str:
    row = grid[r]
    c1 = min(c1, len(row) - 1)
    parts = [row[c] for c in range(max(0, c0), c1 + 1) if row[c]]
    return " | ".join(parts)

def _find_header_cols_in_row_grid(grid: list[list[str]], r: int) -> list[int]:
    """Find columns in row r whose *cell text* contains 'AAA @ BBB'."""
    row = grid[r]
    cols = []
    for c, txt in enumerate(row):
        if txt and ("@" in txt) and _parse_header(txt):
            cols.append(c)
    return cols

def run_matchups(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any]) -> None:
    gb = cfg.get("gameboard")
    if not gb:
        return

    out_rel = (gb.get("out_rel") or "").lstrip(r"\\/") or "data/mlb/latest/matchups"

    # Tunables
    max_scan_rows = int(gb.get("max_scan_rows", 300))
    end_after_blank_rows = int(gb.get("end_after_blank_rows", 8))
    debug = bool(gb.get("debug", False))

    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        # pick dashboard sheet
        want = gb.get("sheet") or ["MLB Game Dashboard", "MLB Dashboard", "Dashboard"]
        want_list = [want] if isinstance(want, str) else list(want)

        def pick_sheet(wants: list[str]) -> Optional[str]:
            if not wants: return None
            lower = {s.lower(): s for s in wb.sheetnames}
            for w in wants:
                if w in wb.sheetnames: return w
                if w.lower() in lower: return lower[w.lower()]
            for w in wants:
                wl = w.lower()
                for s in wb.sheetnames:
                    if wl in s.lower():
                        return s
            return None

        sheet_name = pick_sheet(want_list) or wb.sheetnames[0]
        print(f"‚Ä¢ MLB Matchups (fast): using sheet '{sheet_name}'")
        ws = wb[sheet_name]

        # heuristic for max columns across early rows
        probe_rows = min(40, ws.max_row or 1)
        probe_max_col = 1
        for row in ws.iter_rows(min_row=1, max_row=probe_rows, values_only=True):
            if not row: continue
            last_non_empty = 0
            for idx, val in enumerate(row, start=1):
                if val not in (None, ""):
                    last_non_empty = idx
            probe_max_col = max(probe_max_col, last_non_empty)

        grid = _build_grid(ws, max_rows=max_scan_rows, max_cols=probe_max_col)
        n_rows = len(grid)
        n_cols = probe_max_col

        games: list[dict] = []
        header_hits = 0

        r = 0
        blank_streak = 0
        while r < n_rows:
            header_cols = _find_header_cols_in_row_grid(grid, r)

            if not header_cols:
                if _row_has_any_text(grid, r, 0, n_cols - 1):
                    blank_streak = 0
                else:
                    blank_streak += 1
                    if blank_streak >= end_after_blank_rows and games:
                        if debug:
                            print(f"‚Ä¢ stop at row {r+1}: blank streak {blank_streak}")
                        break
                r += 1
                continue

            header_hits += len(header_cols)
            if debug:
                titles = [grid[r][c] for c in header_cols]
                print(f"  row {r+1} headers: {titles}")

            header_cols_sorted = sorted(header_cols)
            for idx, c_start in enumerate(header_cols_sorted):
                c_end = (header_cols_sorted[idx + 1] - 1) if idx + 1 < len(header_cols_sorted) else (n_cols - 1)

                # Extract and parse "AAA @ BBB"
                title_line = _row_text_slice(grid, r, c_start, c_end)
                parsed = _parse_header(title_line.split("|", 1)[0] if title_line else "")
                if not parsed:
                    continue
                away, home = parsed

                g: Dict[str, Any] = {
                    "away": away, "home": home,
                    "ou": None, "spread_home": None, "ml_home": None, "ml_away": None,
                    "imp_home": None, "imp_away": None,
                    "weather": {"temp_f": None, "wind_mph": None, "desc": None, "is_dome": None},
                    "team_blocks": {
                        "away": {"header": away, "lines": []},
                        "home": {"header": home, "lines": []},
                    },
                }

                # Walk down to find the team bar row
                k = r + 1
                team_bar_row = None
                while k < n_rows:
                    row_slice = grid[k][c_start:c_end+1]
                    if not any(row_slice):
                        k += 1
                        continue
                    left  = next((x for x in row_slice if x), "")
                    right = next((x for x in reversed(row_slice) if x), "")

                    mL = re.match(r"^\s*([A-Z]{2,4})\s*\(([0-9.]+)", left or "")
                    mR = re.match(r"^\s*([A-Z]{2,4})\s*\(([0-9.]+)", right or "")
                    if mL and mR:
                        g["team_blocks"]["away"]["header"] = f"{mL.group(1)} ({mL.group(2)})"
                        g["team_blocks"]["home"]["header"] = f"{mR.group(1)} ({mR.group(2)})"
                        try:
                            g["imp_away"] = float(mL.group(2))
                            g["imp_home"] = float(mR.group(2))
                        except Exception:
                            pass
                        team_bar_row = k
                        break

                    whole = " | ".join([x for x in row_slice if x])
                    U = whole.upper()

                    if "O/U" in U:
                        m_ou = re.search(r"O/?U:\s*([0-9.]+)", whole, flags=re.I)
                        if m_ou: g["ou"] = float(m_ou.group(1))
                        for tm, ml in re.findall(r"\b([A-Z]{2,4})\s*ML:\s*([+-]?\d+)", whole, flags=re.I):
                            if tm.upper() == away: g["ml_away"] = int(ml)
                            if tm.upper() == home: g["ml_home"]  = int(ml)
                    elif "SPREAD" in U:
                        mH = re.search(r"SPREAD:\s*([+-]?[0-9.]+)", whole, flags=re.I)
                        if mH: g["spread_home"] = float(mH.group(1))
                    elif "TOTAL" in U:
                        mA = re.search(rf"{away}\s*([0-9.]+)", whole, flags=re.I)
                        mH = re.search(rf"{home}\s*([0-9.]+)", whole, flags=re.I)
                        if mA: g["imp_away"] = float(mA.group(1))
                        if mH: g["imp_home"]  = float(mH.group(1))
                    elif "WEATHER" in U:
                        is_dome = "DOME" in U
                        g["weather"]["is_dome"] = is_dome
                        g["weather"]["desc"] = None if is_dome else whole.replace("|", " ").strip()

                    k += 1

                if team_bar_row is None:
                    continue

                # Collect player lines until next header in this window or double-blank
                k = team_bar_row + 1
                local_blanks = 0
                while k < n_rows:
                    # stop if a new header appears inside our window
                    row_hdr_cols = [c for c in _find_header_cols_in_row_grid(grid, k) if c_start <= c <= c_end]
                    if row_hdr_cols:
                        break

                    row_slice = grid[k][c_start:c_end+1]
                    left  = next((x for x in row_slice if x), "")
                    right = next((x for x in reversed(row_slice) if x), "")

                    # also stop if team-bar repeats
                    if re.match(r"^\s*[A-Z]{2,4}\s*\([0-9.]+", left or "") and \
                       re.match(r"^\s*[A-Z]{2,4}\s*\([0-9.]+", right or ""):
                        break

                    if not left and not right:
                        local_blanks += 1
                        if local_blanks >= 2:
                            break
                        k += 1
                        continue

                    local_blanks = 0
                    if left:  g["team_blocks"]["away"]["lines"].append(left)
                    if right: g["team_blocks"]["home"]["lines"].append(right)
                    k += 1

                if g.get("ou") is None and all(isinstance(g.get(k2), (int, float)) for k2 in ("imp_home","imp_away")):
                    g["ou"] = float(g["imp_home"]) + float(g["imp_away"])

                games.append(g)

            r += 1

        if debug:
            print(f"‚Ä¢ header candidates seen: {header_hits}")

        out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
        ensure_parent(out_path)
        out_path.write_text(json.dumps(games, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"‚úîÔ∏è  JSON ‚Üí {out_path}  (games: {len(games)})")
        _mark_meta_dir(out_path)
    finally:
        wb.close()


# --------------------------------- config --------------------------------

def _choose_project_root(arg_proj: Optional[str]) -> Path:
    if arg_proj:
        p = Path(arg_proj).resolve()
        if (p / "public").exists():
            return p
        print(f"‚ö†Ô∏è  --project '{p}' has no /public. Falling back to script root.", file=sys.stderr)
    if (ROOT / "public").exists():
        return ROOT
    print("ERROR: could not find a project root with /public.", file=sys.stderr)
    sys.exit(1)


# --------------------------------- main ---------------------------------

def main() -> None:
    print(">>> MLB Exporter (tasks + cheatsheets + matchups)")
    ap = argparse.ArgumentParser(description="Export MLB Excel workbook to site data files (CSV/JSON).")
    ap.add_argument("--xlsm",    default=DEFAULT_XLSM,   help="Path to the source workbook (.xls/.xlsx/.xlsm)")
    ap.add_argument("--project", default=DEFAULT_PROJ,   help="Path to project root (contains /public)")
    ap.add_argument("--config",  default=DEFAULT_CONFIG, help="Path to exporter config JSON")
    args = ap.parse_args()

    xlsm_path     = Path(args.xlsm).resolve()
    project_root  = _choose_project_root(args.project)
    config_path   = Path(args.config).resolve()

    if not xlsm_path.exists():
        print(f"ERROR: workbook not found: {xlsm_path}", file=sys.stderr); sys.exit(1)
    if not config_path.exists():
        print(f"ERROR: config not found: {config_path}", file=sys.stderr); sys.exit(1)

    staged_xlsm, temp_dir = _stage_copy_for_read(xlsm_path)

    try:
        cfg = json.loads(config_path.read_text(encoding="utf-8-sig"))

        tasks = cfg.get("tasks", [])
        if not isinstance(tasks, list):
            print("ERROR: config 'tasks' must be an array.", file=sys.stderr); sys.exit(1)

        for t in tasks:
            sheet = t.get("sheet")
            print(f"\n=== TASK: sheet='{sheet}' | out='{t.get('out_rel','?')}' ===")
            try:
                run_task(staged_xlsm, project_root, t)
            except Exception as e:
                print(f"‚ö†Ô∏è  SKIP: task failed: {e}")

        print("\n=== CHEAT SHEET ===")
        try: run_cheatsheets(staged_xlsm, project_root, cfg)
        except Exception as e: print(f"‚ö†Ô∏è  SKIP cheatsheets: {e}")

        print("\n=== MATCHUPS (MLB Dashboard) ===")
        try: run_matchups(staged_xlsm, project_root, cfg)
        except Exception as e: print(f"‚ö†Ô∏è  SKIP matchups: {e}")

        # finally write meta files for all touched dirs
        print("\n=== META ===")
        _write_meta_files(xlsm_path)

        print("\nDone.")
    finally:
        try: shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception: pass


if __name__ == "__main__":
    main()
