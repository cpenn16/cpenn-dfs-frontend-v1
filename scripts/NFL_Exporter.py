#!/usr/bin/env python3
"""
NFL_Exporter.py — Single meta.json version + separate Player Pool export
-----------------------------------------------------------------------
Exports:
  • tasks (generic sheet → csv/json)
  • cheatsheets (Cheat Sheet)
  • player_pool (Player Pool)   ← NEW, independent of cheatsheets
  • gameboard (NFL Matchups)
  • projections merge with DK/FD salaries + kickoff time

Writes ONE consolidated meta file (configurable via --meta_rel):
  public/data/nfl/classic/latest/meta.json

Usage
-----
python scripts/NFL_Exporter.py --xlsm "C:\\path\\to\\NFL.xlsm" \
                               --project "." \
                               --config "scripts\\configs\\nfl_classic.json"
"""

from __future__ import annotations

import argparse, json, re, sys, shutil, tempfile, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Union

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ------------------------- ROOT / DEFAULT PATHS -------------------------

THIS = Path(__file__).resolve()
ROOT = THIS.parents[1]  # repo root (expected to include /public)

DEFAULT_XLSM     = r"C:\Users\cpenn\Dropbox\Sports Models\NFL\NFL Week 2 Classic.xlsm"
DEFAULT_PROJ     = str(ROOT)
DEFAULT_CONFIG   = str(ROOT / "scripts" / "configs" / "nfl_classic.json")
DEFAULT_META_REL = "data/nfl/classic/latest/meta.json"   # relative to /public

# ------------------------------ Single Meta -----------------------------

class SingleMeta:
    """Collect all artifact stats and write one consolidated meta.json at the end."""
    def __init__(self, project_root: Path, source_workbook: Path, meta_rel: str):
        self.project_root = project_root
        self.source = str(source_workbook)
        self.meta_path = (project_root / "public" / Path(meta_rel)).resolve()
        self._items: List[Dict[str,Any]] = []

    @staticmethod
    def _ts_now_ms() -> int:
        return int(time.time() * 1000)

    @staticmethod
    def _iso_now() -> str:
        return datetime.datetime.now().isoformat(timespec="seconds")

    def add(self, artifact_path: Path, *, sheet: Optional[str]=None,
            record_count: Optional[int]=None, duration_ms: Optional[int]=None,
            tags: Optional[Dict[str, Any]]=None):
        # Store unix-style path under public for consistency
        try:
            rel_under_public = artifact_path.resolve().relative_to((self.project_root / "public").resolve())
            path_str = f"public/{rel_under_public.as_posix()}"
        except Exception:
            path_str = str(artifact_path)

        item = {
            "path": path_str,
            "sheet": sheet,
            "record_count": int(record_count) if record_count is not None else None,
            "duration_ms": int(duration_ms) if duration_ms is not None else None,
        }
        if tags:
            item.update(tags)
        self._items.append(item)

    def flush(self):
        self.meta_path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "last_updated": self._iso_now(),
            "last_updated_ms": self._ts_now_ms(),
            "source_workbook": self.source,
            "artifacts": self._items,
        }
        self.meta_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"📝  META (single) → {self.meta_path}  (items: {len(self._items)})")

# ------------------------------ utilities ------------------------------

def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

def dedup(names: Iterable) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, raw in enumerate(list(names)):
        s = "" if raw is None else str(raw).strip()
        if s == "" or s.lower() in {"nan", "nat"} or s.lower().startswith("unnamed"):
            s = f"col_{i+1}"
        key = s
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        out.append(key)
    return out

def to_json_records(df: pd.DataFrame) -> str:
    df2 = df.astype(object).where(pd.notna(df), "")
    return df2.to_json(orient="records", force_ascii=False, indent=2)

def _stage_copy_for_read(src: Path) -> tuple[Path, Path]:
    """Copy workbook to temp so Excel can stay open while we read."""
    tmpdir = Path(tempfile.mkdtemp(prefix="nfl_export_"))
    dst = tmpdir / src.name
    shutil.copy2(src, dst)
    return dst, tmpdir

# ------------------ generic value helpers (merge uses these) ------------------

def _num_clean(v):
    """Parse numbers that may include $, commas, %, or be blank."""
    if v is None: return None
    s = str(v).replace("$", "").replace(",", "").replace("%", "").strip()
    if s == "": return None
    try: return float(s)
    except Exception: return None

def _first(obj: dict, keys: list[str], default=None):
    for k in keys:
        if k in obj and obj[k] not in (None, ""):
            return obj[k]
    return default

def _key_for(player: str, team: str) -> str:
    p = (player or "").strip().lower()
    t = (team or "").strip().upper()
    return f"{p}|{t}"

def _fmt_money(n):
    if n is None: return ""
    try: return f"{int(round(float(n))):,}"
    except Exception: return ""

# --------------------- openpyxl “display text” formatting --------------------

_PERCENT_RE = re.compile(r"%")

def _decimals_from_format(fmt: str) -> int:
    if not isinstance(fmt, str):
        return 0
    m = re.search(r"0\.([0]+)", fmt)
    return len(m.group(1)) if m else 0

def _format_cell(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    fmt = cell.number_format or ""

    # Dates/times
    if isinstance(v, (datetime.date, datetime.datetime, datetime.time)):
        return str(v)

    # Numbers
    if isinstance(v, (int, float, np.floating)):
        x = float(v)
        if _PERCENT_RE.search(fmt):
            dec = _decimals_from_format(fmt)
            n = x * 100.0 if abs(x) <= 1.01 else x
            if float(n).is_integer():
                return f"{int(round(n))}%"
            return f"{n:.{dec}f}%"
        if float(x).is_integer():
            return str(int(round(x)))
        dec = _decimals_from_format(fmt) or 1
        return f"{x:.{dec}f}"

    return str(v).strip()

# ------------------------------ header normalization --------------------

_HEADER_ALIASES = {
    # Passing
    "pa attempts": "Pa Att",
    "pass attempts": "Pa Att",
    "pa yards": "Pa Yards",
    "pa comp": "Pa Comp",
    "comp %": "Comp%",
    # Rushing
    "ru attempts": "Ru Att",
    "rush attempts": "Ru Att",
    "ru yards": "Ru Yards",
    "rush yards": "Ru Yards",
    "ru td": "Ru TD",
    # Receiving
    "targets": "Targets",
    "tgt share": "Tgt Share",
    "target share": "Tgt Share",
    "rec yards": "Rec Yards",
    # DFS stats
    "dk own%": "DK pOWN%",
    "fd own%": "FD pOWN%",
    "dk pown %": "DK pOWN%",
    "fd pown %": "FD pOWN%",
    "dk opt%": "DK Opt%",
    "fd opt%": "FD Opt%",
    "dk lev%": "DK Lev%",
    "fd lev%": "FD Lev%",
    "dk rtg": "DK Rtg",
    "fd rtg": "FD Rtg",
    # Salaries / team
    "dk sal": "DK Sal",
    "fd sal": "FD Sal",
    "teamabbrev": "Team",
    # Player Pool/MLB-style labels
    "opponent": "Opp",
    "opp": "Opp",
    "o/u": "O/U",
    "imp. total": "Imp. Total",
    "impl. total": "Imp. Total",
    "projection": "Proj",
    "proj": "Proj",
    "value": "Value",
    "pown": "pOWN",
    "pown%": "pOWN",
    "cash/gpp/both": "Cash/GPP/Both",
}

def _norm_header_label(s: str) -> str:
    t = (s or "").replace("\u00A0", " ").replace("\u202F", " ").strip()
    key = re.sub(r"\s+", " ", t).lower()
    return _HEADER_ALIASES.get(key, t)

# ------------------------------ filters engine --------------------------

def _resolve_col(df: pd.DataFrame, name: str) -> Optional[str]:
    if name in df.columns:
        return name
    low_map = {c.lower(): c for c in df.columns}
    return low_map.get((name or "").lower())

def _apply_leaf_filter(df: pd.DataFrame, f: Dict[str, Any]) -> pd.Series:
    col_name = _resolve_col(df, f.get("column", ""))
    if not col_name:
        return pd.Series([True] * len(df), index=df.index)

    op = (f.get("op") or "contains").lower()
    cs = bool(f.get("case_sensitive", False))
    s = df[col_name].astype(str)
    if not cs:
        s = s.str.lower()

    if op == "nonempty":       return s.str.strip().ne("")
    val = str(f.get("value", "")).strip()
    if not cs:                 val = val.lower()

    if   op == "equals":       res = s.eq(val)
    elif op == "not_equals":   res = s.ne(val)
    elif op == "contains":     res = s.str.contains(val, na=False)
    elif op == "not_contains": res = ~s.str.contains(val, na=False)
    elif op == "startswith":   res = s.str.startswith(val, na=False)
    elif op == "endswith":     res = s.str.endswith(val, na=False)
    elif op == "regex":
        try:
            pat = re.compile(val, 0 if cs else re.IGNORECASE)
            res = s.str.match(pat).fillna(False)
        except Exception:
            res = pd.Series([True] * len(df), index=df.index)
    else:
        res = pd.Series([True] * len(df), index=df.index)

    return res.fillna(False)

def _apply_filters(df: pd.DataFrame, filters: Union[List, Dict, None]) -> pd.DataFrame:
    def eval_filter(f) -> pd.Series:
        if isinstance(f, dict) and ("any_of" in f or "all_of" in f):
            if "any_of" in f:
                parts = [eval_filter(x) for x in (f.get("any_of") or [])]
                return pd.concat(parts, axis=1).any(axis=1) if parts else pd.Series([True]*len(df), index=df.index)
            if "all_of" in f:
                parts = [eval_filter(x) for x in (f.get("all_of") or [])]
                return pd.concat(parts, axis=1).all(axis=1) if parts else pd.Series([True]*len(df), index=df.index)
        return _apply_leaf_filter(df, f)

    if not filters: return df
    if isinstance(filters, dict) and ("any_of" in filters or "all_of" in filters):
        return df[eval_filter(filters)]
    if isinstance(filters, list):
        masks = [eval_filter(f) for f in filters]
        return df[pd.concat(masks, axis=1).all(axis=1)] if masks else df
    return df

# ------------------------------ task runner -----------------------------

def maybe_apply_column_mapping(df: pd.DataFrame, mapping: Dict[str, str] | None) -> pd.DataFrame:
    if not mapping: return df
    existing = {src: dst for src, dst in mapping.items() if src in df.columns}
    return df.rename(columns=existing) if existing else df

def reorder_columns_if_all_present(df: pd.DataFrame, order: List[str] | None) -> pd.DataFrame:
    if not order: return df
    return df[order] if all(c in df.columns for c in order) else df

def export_one(df: pd.DataFrame, out_csv: Optional[Path], out_json: Optional[Path], meta: SingleMeta, *, sheet: Optional[str]=None, t0: float=0.0) -> None:
    duration = int((time.time() - t0) * 1000) if t0 else None
    n = int(len(df)) if df is not None else 0
    if out_csv:
        ensure_parent(out_csv)
        df.astype(object).where(pd.notna(df), "").to_csv(out_csv, index=False, encoding="utf-8-sig")
        print(f"✔️  CSV  → {out_csv}")
        meta.add(out_csv, sheet=sheet, record_count=n, duration_ms=duration, tags={"kind":"task","format":"csv"})
    if out_json:
        ensure_parent(out_json)
        out_json.write_text(to_json_records(df), encoding="utf-8")
        print(f"✔️  JSON → {out_json}")
        meta.add(out_json, sheet=sheet, record_count=n, duration_ms=duration, tags={"kind":"task","format":"json"})

def run_task(xlsm_path: Path, project_root: Path, task: Dict[str, Any], meta: SingleMeta) -> None:
    sheet = task.get("sheet")
    if not sheet:
        print("⚠️  SKIP: task missing 'sheet'"); return

    t0 = time.time()
    df = read_literal_table(
        xlsm_path=xlsm_path,
        sheet=sheet,
        header_row=task.get("header_row"),
        data_start_row=task.get("data_start_row"),
        limit_to_col=task.get("limit_to_col"),
    )

    keep_cols_src: List[str] = task.get("keep_columns_sheet_order", [])
    if keep_cols_src:
        df = df[[c for c in df.columns if c in keep_cols_src]]

    df = maybe_apply_column_mapping(df, task.get("column_mapping"))
    df = reorder_columns_if_all_present(df, task.get("column_order"))
    df = _apply_filters(df, task.get("filters"))

    out_rel = (task.get("out_rel") or "").lstrip(r"\/")
    if not out_rel:
        print(f"⚠️  SKIP: task for '{sheet}' missing 'out_rel'"); return

    base = project_root / "public" / Path(out_rel)
    fmt = str(task.get("format", "json")).lower()
    csv_path  = base.with_suffix(".csv")  if fmt in ("csv", "both")  else None
    json_path = base.with_suffix(".json") if fmt in ("json", "both") else None
    export_one(df, csv_path, json_path, meta, sheet=sheet, t0=t0)

# ----------------------- header-on-next-row helper ----------------------

_HEADER_KEYS = {"player","salary","team","opponent","opp","o/u","imp. total","proj","projection","value","pown","pown%","cash/gpp/both","time"}

def _looks_like_header(vals: List[str]) -> bool:
    tokens = {re.sub(r"\s+"," ",str(v or "")).strip().lower() for v in vals}
    return len(tokens & _HEADER_KEYS) >= 2

def _maybe_shift_header_down(ws: Worksheet, header_r: int, start_c: int, width: int, n_cols: int, title_text: str) -> int:
    """
    If the current header row is the section title (e.g., 'Cash Core') and the *next* row
    looks like real headers (Player, Salary, Team, ...), shift header down by +1.
    """
    head_vals = [_format_cell(ws.cell(header_r, c)) for c in range(start_c, min(start_c+width, n_cols+1))]
    first_cell = (head_vals[0] or "").strip()
    if first_cell == (title_text or "").strip():
        nxt = header_r + 1
        nxt_vals = [_format_cell(ws.cell(nxt, c)) for c in range(start_c, min(start_c+width, n_cols+1))]
        if _looks_like_header(nxt_vals):
            return nxt
    return header_r

# -------------------------- cheatsheets (by title) ----------------------

def run_cheatsheets(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any], meta: SingleMeta) -> None:
    cs = cfg.get("cheatsheets")
    if not cs: return
    sheet      = cs.get("sheet") or "Cheat Sheet"
    out_rel    = (cs.get("out_rel") or "").lstrip(r"\/")
    title_ci   = bool(cs.get("title_match_ci", True))
    limit_rows = int(cs.get("limit_rows", 10))
    if not out_rel:
        print("⚠️  SKIP cheatsheets: missing out_rel"); return

    t0_total = time.time()
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            print(f"⚠️  SKIP cheatsheets: sheet '{sheet}' not found"); return
        ws = wb[sheet]
        n_rows, n_cols = ws.max_row, ws.max_column

        def norm(s: Any) -> str:
            txt = "" if s is None else str(s).strip()
            return txt.lower() if title_ci else txt

        titles_cfg = cs.get("tables") or []
        all_titles_norm = {norm(t.get("title")) for t in titles_cfg if t.get("title")}

        # Fast index of first occurrences of every non-empty cell text
        index: Dict[str, tuple] = {}
        max_scan_rows = min(n_rows, int(cs.get("max_scan_rows", n_rows)))
        for r in range(1, max_scan_rows + 1):
            row = ws[r]
            for c in range(1, min(n_cols, len(row)) + 1):
                s = norm(row[c-1].value)
                if s and s not in index:
                    index[s] = (r, c)

        tables_out: List[Dict[str, Any]] = []
        for i, t in enumerate(titles_cfg):
            t0 = time.time()
            title = str(t.get("title") or f"Table {i+1}").strip()
            width = max(1, int(t.get("width", 3)))
            loc = index.get(norm(title))
            if not loc:
                print(f"⚠️  cheatsheets: title not found: '{title}'")
                continue
            start_r, start_c = loc

            # ← FIX: if current row is the section title, push header row down one line when needed
            header_r = _maybe_shift_header_down(ws, start_r, start_c, width, n_cols, title)
            data_r0  = header_r + 1

            hdr_cells = [ws.cell(header_r, c) for c in range(start_c, min(start_c + width, n_cols + 1))]
            headers = dedup([_norm_header_label(_format_cell(c)) for c in hdr_cells])

            rows = []
            r = data_r0
            blank_rows = 0
            while r <= n_rows and len(rows) < limit_rows:
                row_cells = [ws.cell(r, c) for c in range(start_c, start_c + len(headers))]
                display = [_format_cell(c) for c in row_cells]
                if all(x == "" for x in display):
                    blank_rows += 1
                    if blank_rows >= 2:
                        break
                    r += 1
                    continue
                blank_rows = 0
                if norm(ws.cell(r, start_c).value) in all_titles_norm:
                    break
                rows.append(display)
                r += 1

            sub = pd.DataFrame(rows, columns=headers)

            # Normalize a "Player" column for Player Pool/pos tables
            if "Player" not in sub.columns:
                for cand in ["QB","RB","WR","TE","Name","PLAYER"]:
                    if cand in sub.columns:
                        sub = sub.rename(columns={cand:"Player"})
                        break

            tables_out.append({
                "id":      f"t{i+1}",
                "label":   title,
                "columns": list(sub.columns),
                "rows":    sub.astype(object).where(pd.notna(sub), "").to_dict(orient="records"),
            })
            print(f"• table '{title}' rows={len(sub)} in {int((time.time()-t0)*1000)} ms")

        out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
        ensure_parent(out_path)
        out_path.write_text(json.dumps({"tables": tables_out}, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✔️  JSON → {out_path}  (tables written: {len(tables_out)} of {len(titles_cfg)})")
        meta.add(out_path, sheet=sheet, record_count=sum(len(t['rows']) for t in tables_out),
                 duration_ms=int((time.time()-t0_total)*1000), tags={"kind":"cheatsheets"})
    finally:
        wb.close()

# ------------------------------- literal read ---------------------------

def _excel_col_to_idx(label: str) -> int:
    s = re.sub(r"[^A-Za-z]", "", str(label)).upper()
    if not s: return 0
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1

def read_literal_table(xlsm_path: Path, sheet: str,
                       header_row: Optional[int],
                       data_start_row: Optional[int],
                       limit_to_col: Optional[str] = None) -> pd.DataFrame:
    """
    Read a sheet using openpyxl and return a DataFrame of *strings* matching Excel display.
    `limit_to_col` (e.g., "AE") caps the rightmost column read.
    """
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        ws = wb[sheet]

        max_c = ws.max_column
        if limit_to_col:
            try:
                max_c = min(max_c, _excel_col_to_idx(limit_to_col) + 1)
            except Exception:
                pass

        if header_row is None or data_start_row is None:
            scan = min(8, ws.max_row)
            best_r, best_nonempty = 1, -1
            for r in range(1, scan + 1):
                vals = [c.value for c in ws[r][0:max_c]]
                nonempty = sum(1 for x in vals if x not in (None, ""))
                if nonempty > best_nonempty:
                    best_nonempty = nonempty
                    best_r = r
            header_row = best_r
            data_start_row = best_r + 1

        raw_headers = [_format_cell(c) for c in ws[header_row][0:max_c]]
        raw_headers = [_norm_header_label(h) for h in raw_headers]
        headers = dedup(raw_headers)

        out_rows: List[List[str]] = []
        blanks_in_a_row = 0
        for r in range(int(data_start_row), ws.max_row + 1):
            cells = ws[r][0:max_c]
            row = [_format_cell(c) for c in cells]
            if all(v == "" for v in row):
                blanks_in_a_row += 1
                if blanks_in_a_row >= 3: break
                continue
            blanks_in_a_row = 0
            out_rows.append(row)

        df = pd.DataFrame(out_rows, columns=headers)
        df = df.dropna(axis=0, how="all")
        df = df.replace("", np.nan).dropna(axis=0, how="all").fillna("")
        df = df.loc[:, ~(df.astype(str).eq("").all())]
        return df
    finally:
        wb.close()

# ---------------------- NFL GAMEBOARD (Dashboard) -----------------------

def _gb_cell(ws: Worksheet, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()

def _gb_row_text_range(ws: Worksheet, r: int, c0: int, c1: int) -> str:
    """Join non-empty cell texts inside [c0..c1] for row r."""
    parts = [_gb_cell(ws, r, c) for c in range(max(1, c0), max(1, c1) + 1)]
    parts = [p for p in parts if p]
    return " | ".join(parts)

_TEAM_BAR_RE = re.compile(r"^\s*([A-Z]{2,4})\s*\(([-+]?[0-9.]+)\)\s*$")

def _parse_team_bar(txt: str):
    m = _TEAM_BAR_RE.match(txt or "")
    if not m:
        return None
    return m.group(1).upper(), float(m.group(2))

def _gb_find_header_cols_in_row(ws: Worksheet, r: int, max_col: int,
                                yellow_rgbs: set, title_re: re.Pattern,
                                _cache: dict=None) -> list[int]:
    cols = []
    for c in range(1, max_col + 1):
        key = (r, c)
        txt = ws.cell(r, c).value
        if not txt:
            continue

        is_header = False
        try:
            if _cache is not None and key in _cache:
                rgb = _cache[key]
            else:
                fill = ws.cell(r, c).fill
                rgb = (fill.fgColor.rgb or "").upper() if (fill and fill.patternType == "solid") else ""
                if _cache is not None:
                    _cache[key] = rgb
            if rgb in yellow_rgbs:
                is_header = True
        except Exception:
            pass
        if isinstance(txt, str) and title_re.match(txt.strip()):
            is_header = True

        if is_header:
            cols.append(c)
    return cols

def _gb_parse_ml_pieces(s: str) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for tm, ml in re.findall(r"\b([A-Z]{2,4})\s*ML:\s*([+-]?\d+)", s, flags=re.I):
        out[tm.upper()] = int(ml)
    return out

def _gb_parse_spread_pieces(s: str) -> Dict[str, float]:
    m = re.search(r"Spread:\s*([A-Z]{2,4})\s*([+-]?[0.9]+)\s*\|\s*([A-Z]{2,4})\s*([+-]?[0-9.]+)", s, flags=re.I)
    if not m: return {}
    return {m.group(1).upper(): float(m.group(2)), m.group(3).upper(): float(m.group(4))}

def _gb_parse_totals_pieces(s: str) -> Dict[str, float]:
    m = re.search(r"Totals?:\s*([A-Z]{2,4})\s*([0-9.]+)\s*\|\s*([A-Z]{2,4})\s*([0-9.]+)", s, flags=re.I)
    if not m: return {}
    return {m.group(1).upper(): float(m.group(2)), m.group(3).upper(): float(m.group(4))}

def _gb_parse_weather(s: str) -> Dict[str, Any]:
    is_dome = "dome" in s.lower()
    temp = re.search(r"([0-9.]+)\s*°?F", s, flags=re.I)
    wind = re.search(r"([0-9.]+)\s*mph", s, flags=re.I)
    return {
        "temp_f": float(temp.group(1)) if temp else None,
        "wind_mph": float(wind.group(1)) if wind else None,
        "desc": None if is_dome else s.replace("|", " ").strip(),
        "is_dome": is_dome,
    }

def _pick_dashboard_sheet(wb, sheet_cfg):
    want_list = [sheet_cfg] if isinstance(sheet_cfg, str) else list(sheet_cfg or [])
    if not want_list:
        return None

    lower_map = {s.lower(): s for s in wb.sheetnames}
    for want in want_list:
        nm = str(want).strip()
        if nm in wb.sheetnames:
            return nm
        if nm.lower() in lower_map:
            return lower_map[nm.lower()]

    for want in want_list:
        w = str(want).lower()
        for actual in wb.sheetnames:
            if w in actual.lower():
                return actual
    return None

def run_gameboard(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any], meta: SingleMeta) -> None:
    gb = cfg.get("gameboard")
    if not gb: return

    out_rel = (gb.get("out_rel") or "").lstrip(r"\\/")
    if not out_rel:
        print("⚠️  SKIP gameboard: missing 'out_rel'"); return

    yellow_rgbs = {str(x).upper() for x in gb.get("header_yellow_rgb", ["FFFFE699","FFFFF2CC","FFFFFF00"])}
    title_re = re.compile(gb.get("title_regex", r"^[A-Z]{2,4}\s*@\s*[A-Z]{2,4}$"))

    t0_total = time.time()
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        sheet_name = _pick_dashboard_sheet(wb, gb.get("sheet", "NFL Game Dashboard"))
        if not sheet_name:
            print("⚠️  SKIP gameboard: dashboard sheet not found"); return
        print(f"• Gameboard: using sheet '{sheet_name}'")
        ws = wb[sheet_name]
        max_row, max_col = ws.max_row, ws.max_column
        color_cache = {}

        games: List[Dict[str, Any]] = []

        r = 1
        while r <= max_row:
            header_cols = _gb_find_header_cols_in_row(ws, r, max_col, yellow_rgbs, title_re, _cache=color_cache)
            if not header_cols:
                r += 1
                continue

            header_cols_sorted = sorted(header_cols)
            for idx, c_start in enumerate(header_cols_sorted):
                c_end = (header_cols_sorted[idx + 1] - 1) if idx + 1 < len(header_cols_sorted) else max_col

                title_line = _gb_row_text_range(ws, r, c_start, c_end)
                title = (title_line.split("|", 1)[0] or "").strip()
                m_title = title_re.match(title)
                away, home = (m_title.group(1), m_title.group(2)) if m_title else ("", "")

                g: Dict[str, Any] = {
                    "date": None,
                    "away": away,
                    "home": home,
                    "ou": None,
                    "spread_home": None,
                    "ml_home": None,
                    "ml_away": None,
                    "weather": None,
                    "imp_home": None,
                    "imp_away": None,
                    "team_blocks": {
                        "away": {"header": "", "lines": []},
                        "home": {"header": "", "lines": []},
                    },
                }

                k = r + 1
                team_bar_row = None
                blank_guard = 0
                while k <= max_row:
                    vals = [_gb_cell(ws, k, c) for c in range(c_start, c_end + 1)]
                    left  = next((x for x in vals if x), "")
                    right = next((x for x in reversed(vals) if x), "")

                    if not (left or right):
                        blank_guard += 1
                        if blank_guard >= 2:
                            break
                        k += 1
                        continue

                    mL = _parse_team_bar(left)
                    mR = _parse_team_bar(right)
                    if mL and mR:
                        g["team_blocks"]["away"]["header"] = left
                        g["team_blocks"]["home"]["header"] = right
                        if not g["away"]: g["away"] = mL[0]
                        if not g["home"]: g["home"] = mR[0]
                        if g.get("imp_away") is None: g["imp_away"] = mL[1]
                        if g.get("imp_home")  is None: g["imp_home"]  = mR[1]
                        team_bar_row = k
                        break

                    whole = " | ".join([x for x in vals if x])
                    U = whole.upper()
                    if "O/U" in U:
                        m_ou = re.search(r"O/?U:\s*([0-9.]+)", whole, flags=re.I)
                        if m_ou: g["ou"] = float(m_ou.group(1))
                        ml = _gb_parse_ml_pieces(whole)
                        if g["away"] in ml: g["ml_away"] = ml[g["away"]]
                        if g["home"] in ml: g["ml_home"]  = ml[g["home"]]
                    elif "SPREAD" in U:
                        sp = _gb_parse_spread_pieces(whole)
                        if g["home"] in sp: g["spread_home"] = sp[g["home"]]
                    elif "TOTAL" in U:
                        tp = _gb_parse_totals_pieces(whole)
                        if g["away"] in tp: g["imp_away"] = float(tp[g["away"]])
                        if g["home"] in tp: g["imp_home"]  = float(tp[g["home"]])
                    elif "WEATHER" in U:
                        g["weather"] = _gb_parse_weather(whole.split(":", 1)[1] if ":" in whole else whole)

                    k += 1

                if not team_bar_row:
                    continue

                k = team_bar_row + 1
                blank_rows = 0
                while k <= max_row:
                    row_hdr_cols = _gb_find_header_cols_in_row(ws, k, c_end, yellow_rgbs, title_re, _cache=color_cache)
                    row_hdr_cols = [c for c in row_hdr_cols if c_start <= c <= c_end]
                    if row_hdr_cols:
                        break

                    vals = [_gb_cell(ws, k, c) for c in range(c_start, c_end + 1)]
                    left  = next((x for x in vals if x), "")
                    right = next((x for x in reversed(vals) if x), "")

                    if _parse_team_bar(left) and _parse_team_bar(right):
                        break

                    if not left and not right:
                        blank_rows += 1
                        if blank_rows >= 2:
                            break
                        k += 1
                        continue

                    blank_rows = 0
                    if left:  g["team_blocks"]["away"]["lines"].append(left)
                    if right: g["team_blocks"]["home"]["lines"].append(right)
                    k += 1

                games.append(g)

            r += 1

        for g in games:
            if g.get("ou") is None and all(isinstance(g.get(k), (int, float)) for k in ("imp_home", "imp_away")):
                g["ou"] = float(g["imp_home"]) + float(g["imp_away"])
            if g.get("imp_home") is None and isinstance(g.get("ou"), (int, float)) and isinstance(g.get("spread_home"), (int, float)):
                ou = float(g["ou"]); s = float(g["spread_home"])
                g["imp_home"] = (ou - s) / 2
                g["imp_away"] = ou - g["imp_home"]

        out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
        ensure_parent(out_path)
        out_path.write_text(json.dumps(games, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✔️  JSON → {out_path}  (games: {len(games)})")
        meta.add(out_path, sheet=sheet_name, record_count=len(games),
                 duration_ms=int((time.time()-t0_total)*1000), tags={"kind":"gameboard"})
    finally:
        wb.close()

# ------------------------- JSON helpers (once) --------------------------

def _load_json(path: Path):
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except FileNotFoundError:
        return None

def _save_json(path: Path, obj):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def _to_rows_shape(raw):
    if raw is None:
        return [], ("none", None)
    if isinstance(raw, list):
        return raw, ("array", None)
    if isinstance(raw, dict):
        if isinstance(raw.get("rows"), list):
            return raw["rows"], ("rows", raw)
        if isinstance(raw.get("players"), list):
            return raw["players"], ("players", raw)
    return [], ("unknown", raw)

def _write_back(rows, shape, out_path: Path):
    kind, container = shape
    if kind == "array":
        _save_json(out_path, rows)
    elif kind == "rows":
        container["rows"] = rows
        _save_json(out_path, container)
    elif kind == "players":
        container["players"] = rows
        _save_json(out_path, container)
    else:
        _save_json(out_path, rows)

# ------------------------- SALARY MERGE (with TIME) ---------------------

_TIME_RE = re.compile(r"(\d{1,2})\s*:\s*(\d{2})(?::\d{2})?\s*([ap])\s*\.?\s*m", re.I)

def _normalize_time_string(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    m = _TIME_RE.search(str(s))
    if not m:
        return None
    h = int(m.group(1))
    mm = m.group(2).zfill(2)
    ampm = "AM" if m.group(3).lower() == "a" else "PM"
    return f"{h}:{mm} {ampm}"

def _time_from_game_info(gi: Optional[str]) -> Optional[str]:
    if not gi:
        return None
    m = re.search(r"(\d{1,2}\s*:\s*\d{2}\s*[ap]\s*\.?\s*m)", str(gi), flags=re.I)
    return _normalize_time_string(m.group(1)) if m else None

def _pick_sheet_ci(wb, want_list: list[str]) -> Optional[str]:
    if not want_list: return None
    lower_map = {s.lower(): s for s in wb.sheetnames}
    for w in want_list:
        if w in wb.sheetnames: return w
        if w.lower() in lower_map: return lower_map[w.lower()]
    for s in wb.sheetnames:
        sl = s.lower()
        if ("dk" in sl or "draft" in sl) and ("sal" in sl or "salar" in sl):
            return s
    return None

_NAME_WITH_ID_RE = re.compile(r"\s*(.+?)\s*\(\d+\)\s*$")

def _name_from_name_plus_id(v: str) -> str:
    if not v: return ""
    m = _NAME_WITH_ID_RE.match(str(v))
    return (m.group(1) if m else str(v)).strip()

def _build_kickoff_map_from_workbook(xlsm_path: Path) -> dict[str, str]:
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        sheet_name = _pick_sheet_ci(wb, ["DK Salaries", "DraftKings Salaries", "Salaries"])
        if not sheet_name:
            return {}
        df = read_literal_table(xlsm_path, sheet_name, header_row=None, data_start_row=None, limit_to_col=None)

        def col(*names):
            for n in names:
                c = _resolve_col(df, n)
                if c: return c
            return None

        name_col  = col("Name + ID", "Name", "Player")
        team_col  = col("TeamAbbrev", "Team")
        time_col  = col("Time ET", "Time", "StartTime", "Start Time", "Kickoff", "Column1.3")
        gi_col    = col("Game Info", "GameInfo", "Column1", "Column1.1")

        if not name_col or not team_col:
            return {}

        kick: dict[str, str] = {}
        for _, row in df.iterrows():
            raw_name = str(row.get(name_col, "")).strip()
            player   = _name_from_name_plus_id(raw_name)
            team     = str(row.get(team_col, "")).strip().upper()
            if not player or not team:
                continue

            time_raw = str(row.get(time_col, "")).strip() if time_col else ""
            t = _normalize_time_string(time_raw) if time_raw else None
            if not t and gi_col:
                t = _time_from_game_info(str(row.get(gi_col, "")).strip())

            if t:
                kick[_key_for(player, team)] = t
        return kick
    finally:
        wb.close()

def _build_salary_map(base_dir: Path):
    files = ["qb_data.json","rb_data.json","wr_data.json","te_data.json","dst_data.json"]
    m = {}
    seen = 0
    for fn in files:
        p = base_dir / fn
        arr = _load_json(p)
        if not isinstance(arr, list):
            continue
        for r in arr:
            player = _first(r, ["player","Player","name","Name"])
            team   = _first(r, ["team","Team","Tm","TEAM","TeamAbbrev","teamabbrev"])
            if not player or not team:
                continue

            dk_sal = _num_clean(_first(r, ["DK Sal","dk_sal","DK_Sal","DKSalary","DK Salary","salary_dk","dk"]))
            fd_sal = _num_clean(_first(r, ["FD Sal","fd_sal","FD_Sal","FDSalary","FD Salary","salary_fd","fd"]))

            dk_time_raw = _first(r, [
                "Time", "Time ET", "TimeET", "StartTime", "Start Time", "Kickoff",
                "Column1.3",
            ])
            if not dk_time_raw:
                dk_time_raw = _time_from_game_info(_first(r, ["Game Info","GameInfo","Column1","Column1.1"]))
            time_str = _normalize_time_string(dk_time_raw) if dk_time_raw else None

            if dk_sal is None and fd_sal is None and not time_str:
                continue

            m[_key_for(player, team)] = {"dk": dk_sal, "fd": fd_sal, "time": time_str}
            seen += 1
    return m, seen

def merge_salaries_into_projections(project_root: Path,
                                    base_rel="data/nfl/classic/latest",
                                    kickoff_map: Optional[dict[str, str]] = None,
                                    meta: Optional[SingleMeta] = None) -> None:
    base = project_root / "public" / base_rel
    proj_path = base / "projections.json"
    if not proj_path.exists():
        print(f"⚠️  merge_salaries: {proj_path} not found; skipping")
        return

    t0 = time.time()
    raw = _load_json(proj_path)
    rows, shape = _to_rows_shape(raw)

    sal_map, seen_src = _build_salary_map(base)

    updated, dk_hits, fd_hits, time_hits_json, time_hits_sheet = 0, 0, 0, 0, 0
    for r in rows:
        player = _first(r, ["player","Player","name","Name"], "")
        team   = _first(r, ["team","Team","Tm","TEAM","TeamAbbrev","teamabbrev"], "")
        key    = _key_for(player, team)
        s = sal_map.get(key)

        if s:
            if s.get("dk") is not None:
                r["dk_sal"] = float(s["dk"])
                r["DK Sal"] = _fmt_money(s["dk"])
                dk_hits += 1
            if s.get("fd") is not None:
                r["fd_sal"] = float(s["fd"])
                r["FD Sal"] = _fmt_money(s["fd"])
                fd_hits += 1
            if s.get("time") and not r.get("time"):
                r["time"] = s["time"]
                time_hits_json += 1

        if not r.get("time") and kickoff_map:
            t2 = kickoff_map.get(key)
            if t2:
                r["time"] = t2
                time_hits_sheet += 1

        if s or (kickoff_map and key in kickoff_map):
            updated += 1

    _write_back(rows, shape, proj_path)

    print("\n=== MERGE SALARIES INTO projections.json ===")
    print(f"• Source salary rows seen: {seen_src:,}")
    print(f"• Projections updated:     {updated:,}")
    print(f"• DK hits:                 {dk_hits:,}")
    print(f"• FD hits:                 {fd_hits:,}")
    print(f"• Time hits (json):        {time_hits_json:,}")
    print(f"• Time hits (DK sheet):    {time_hits_sheet:,}")
    print(f"• Output:                  {proj_path}")
    if meta:
        meta.add(proj_path, sheet=None, record_count=len(rows),
                 duration_ms=int((time.time()-t0)*1000), tags={"kind":"merge","dk_hits":dk_hits,"fd_hits":fd_hits})

# ---------------------- Player Pool enrich (name/role/time) --------------

def _enrich_player_pool_json(out_path: Path, xlsm_path: Path) -> None:
    pp = _load_json(out_path)
    if not isinstance(pp, dict) or "tables" not in pp:
        return

    kickoff_map = _build_kickoff_map_from_workbook(xlsm_path)

    def pick_player(d: dict) -> str:
        return _first(d, ["Player","QB","RB","WR","TE","Name","PLAYER"], "")

    def pick_team(d: dict) -> str:
        return _first(d, ["Team","Tm","TEAM","TeamAbbrev","teamabbrev"], "")

    for tbl in pp.get("tables", []):
        label = (tbl.get("label") or "").strip()
        rows  = tbl.get("rows") or []

        # Default Cash/GPP/Both for Cash Core
        if label.lower() == "cash core":
            for r in rows:
                if not _first(r, ["Cash/GPP/Both","Role"], None):
                    r["Cash/GPP/Both"] = "Cash"

        # Ensure 'Player' exists (normalize QB/RB/WR/TE header cases)
        if rows and "Player" not in (tbl.get("columns") or []):
            for r in rows:
                p = pick_player(r)
                if p and "Player" not in r:
                    r["Player"] = p
            cols = list(tbl.get("columns") or [])
            if "Player" not in cols:
                cols.insert(0, "Player")
            tbl["columns"] = cols

        # Fill missing Time via kickoff_map
        for r in rows:
            if not r.get("Time") and not r.get("time"):
                key = _key_for(pick_player(r), pick_team(r).upper())
                t = kickoff_map.get(key)
                if t:
                    r["Time"] = t

    _save_json(out_path, pp)

# ---------------------------- optional Player Pool ----------------------

def _run_optional_player_pool(staged_xlsm: Path, project_root: Path, cfg: dict, meta: SingleMeta) -> None:
    """
    Reuse run_cheatsheets() for a second, independent export driven by cfg['player_pool'].
    Then enrich the JSON: ensure Player col, default role for Cash Core, add Time from DK sheet.
    """
    pp = cfg.get("player_pool")
    if not isinstance(pp, dict):
        return
    print("• Player Pool: exporting from sheet:", pp.get("sheet", "Player Pool"))
    run_cheatsheets(staged_xlsm, project_root, {"cheatsheets": pp}, meta)

    out_rel = (pp.get("out_rel") or "").lstrip(r"\/")
    if out_rel:
        out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
        try:
            _enrich_player_pool_json(out_path, staged_xlsm)
            print(f"• Player Pool enriched (name/role/time) → {out_path}")
        except Exception as e:
            print(f"⚠️  Player Pool enrich failed: {e}")

# --------------------------------- main ---------------------------------

def _choose_project_root(arg_proj: Optional[str]) -> Path:
    if arg_proj:
        p = Path(arg_proj).resolve()
        if (p / "public").exists():
            return p
        print(f"⚠️  --project '{p}' has no /public. Falling back to script root.", file=sys.stderr)
    if (ROOT / "public").exists():
        return ROOT
    print("ERROR: could not find a project root with /public.", file=sys.stderr)
    sys.exit(1)

def main() -> None:
    print(">>> NFL Exporter (single meta.json + player_pool)")
    ap = argparse.ArgumentParser(description="Export NFL Excel workbook to site data files (CSV/JSON).")
    ap.add_argument("--xlsm",    default=DEFAULT_XLSM,   help="Path to the source workbook (.xls/.xlsx/.xlsm)")
    ap.add_argument("--project", default=DEFAULT_PROJ,   help="Path to project root (contains /public)")
    ap.add_argument("--config",  default=DEFAULT_CONFIG, help="Path to exporter config JSON")
    ap.add_argument("--meta_rel", default=DEFAULT_META_REL, help="Relative path under /public for consolidated meta.json")
    args = ap.parse_args()

    xlsm_path     = Path(args.xlsm).resolve()
    project_root  = _choose_project_root(args.project)
    config_path   = Path(args.config).resolve()

    if not xlsm_path.exists():
        print(f"ERROR: workbook not found: {xlsm_path}", file=sys.stderr); sys.exit(1)
    if not config_path.exists():
        print(f"ERROR: config not found: {config_path}", file=sys.stderr); sys.exit(1)

    staged_xlsm, temp_dir = _stage_copy_for_read(xlsm_path)
    meta = SingleMeta(project_root=project_root, source_workbook=xlsm_path, meta_rel=args.meta_rel)

    try:
        cfg = json.loads(config_path.read_text(encoding="utf-8-sig"))

        tasks = cfg.get("tasks", [])
        if not isinstance(tasks, list):
            print("ERROR: config 'tasks' must be an array.", file=sys.stderr); sys.exit(1)

        for t in tasks:
            sheet = t.get("sheet")
            print(f"\n=== TASK: sheet='{sheet}' | out='{t.get('out_rel','?')}' ===")
            try:
                run_task(staged_xlsm, project_root, t, meta)
            except Exception as e:
                print(f"⚠️  SKIP: task failed: {e}")

        print("\n=== CHEAT SHEETS ===")
        try:
            # existing Cheat Sheet export
            run_cheatsheets(staged_xlsm, project_root, cfg, meta)
        except Exception as e:
            print(f"⚠️  SKIP cheatsheets: {e}")

        print("\n=== PLAYER POOL ===")
        try:
            _run_optional_player_pool(staged_xlsm, project_root, cfg, meta)
        except Exception as e:
            print(f"⚠️  SKIP player pool: {e}")

        print("\n=== GAMEBOARD (NFL Matchups) ===")
        try:
            run_gameboard(staged_xlsm, project_root, cfg, meta)
        except Exception as e:
            print(f"⚠️  SKIP gameboard: {e}")

        # Build kickoff map once for projections merge
        try:
            kickoff_map = _build_kickoff_map_from_workbook(staged_xlsm)
        except Exception as e:
            print(f"⚠️  Building kickoff map failed: {e}")
            kickoff_map = {}

        try:
            merge_salaries_into_projections(project_root, "data/nfl/classic/latest", kickoff_map=kickoff_map, meta=meta)
        except Exception as e:
            print(f"⚠️  Salary merge failed: {e}")

        # one write at the end
        meta.flush()

        print("\nDone.")
    finally:
        try: shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception: pass

if __name__ == "__main__":
    main()
