#!/usr/bin/env python3
"""
NFL_WeeklyIdsAndXwalk.py — build DK/FD site IDs and the name crosswalk
EVERY time you run it. No freshness guard, no skipping.

Usage examples:
  python scripts/NFL_WeeklyIdsAndXwalk.py
  python scripts/NFL_WeeklyIdsAndXwalk.py --xlsm "C:\\path\\NFL Week 1 Classic.xlsm"
  python scripts/NFL_WeeklyIdsAndXwalk.py --project "." --config scripts\\configs\\nfl_classic.json
  python scripts/NFL_WeeklyIdsAndXwalk.py --only-site-ids
  python scripts/NFL_WeeklyIdsAndXwalk.py --only-xwalk
"""

from __future__ import annotations

import argparse, json, re, sys, shutil, tempfile, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
import numpy as np
from openpyxl import load_workbook

# ------------------------- defaults/paths -------------------------

THIS = Path(__file__).resolve()
ROOT = THIS.parents[1]  # repo root (should contain /public)

DEFAULT_XLSM   = r"C:\Users\cpenn\Dropbox\Sports Models\NFL\NFL Week 2 Classic.xlsm"
DEFAULT_PROJ   = str(ROOT)
DEFAULT_CONFIG = str(ROOT / "scripts" / "configs" / "nfl_classic.json")

# ------------------------- small utils ---------------------------

def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

def dedup(names) -> List[str]:
    seen, out = {}, []
    for i, raw in enumerate(list(names)):
        s = "" if raw is None else str(raw).strip()
        if s == "" or s.lower().startswith("unnamed"):
            s = f"col_{i+1}"
        key = s
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        out.append(key)
    return out

def to_json_records(df: pd.DataFrame) -> str:
    df2 = df.astype(object).where(pd.notna(df), "")
    return df2.to_json(orient="records", force_ascii=False, indent=2)

def _stage_copy_for_read(src: Path) -> tuple[Path, Path]:
    """Copy workbook to temp so it can stay open in Excel while we read."""
    tmpdir = Path(tempfile.mkdtemp(prefix="nfl_weekly_"))
    dst = tmpdir / src.name
    shutil.copy2(src, dst)
    return dst, tmpdir

# ------------------------ Excel display helpers -------------------

_PERCENT_RE = re.compile(r"%")

def _decimals_from_format(fmt: str) -> int:
    if not isinstance(fmt, str): return 0
    m = re.search(r"0\.([0]+)", fmt)
    return len(m.group(1)) if m else 0

def _format_cell(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    fmt = cell.number_format or ""

    # datetimes: let openpyxl give us python objects; stringify
    if isinstance(v, (datetime.date, datetime.datetime, datetime.time)):
        return str(v)

    if isinstance(v, (int, float, np.floating)):
        x = float(v)
        if _PERCENT_RE.search(fmt):
            dec = _decimals_from_format(fmt)
            n = x * 100.0 if abs(x) <= 1.01 else x
            if float(n).is_integer():
                return f"{int(round(n))}%"
            return f"{n:.{dec}f}%"
        if float(x).is_integer():
            return str(int(round(x)))
        dec = _decimals_from_format(fmt) or 1
        return f"{x:.{dec}f}"

    return str(v).strip()

# ------------------------ header normalization --------------------

_HEADER_ALIASES = {
    "teamabbrev": "Team",
    "dk sal": "DK Sal",
    "fd sal": "FD Sal",
    "dk pown %": "DK pOWN%",
    "fd pown %": "FD pOWN%",
}
def _norm_header_label(s: str) -> str:
    t = (s or "").replace("\u00A0", " ").replace("\u202F", " ").strip()
    key = re.sub(r"\s+", " ", t).lower()
    return _HEADER_ALIASES.get(key, t)

def _resolve_col(df: pd.DataFrame, name: str) -> Optional[str]:
    if name in df.columns: return name
    low = {c.lower(): c for c in df.columns}
    return low.get((name or "").lower())

# ------------------------ literal table reader --------------------

def _excel_col_to_idx(label: str) -> int:
    s = re.sub(r"[^A-Za-z]", "", str(label)).upper()
    if not s: return 0
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1

def read_literal_table(xlsm_path: Path, sheet: str,
                       header_row: Optional[int],
                       data_start_row: Optional[int],
                       limit_to_col: Optional[str] = None) -> pd.DataFrame:
    """
    Read a sheet and return a DataFrame of strings that match what Excel shows.
    """
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        ws = wb[sheet]

        max_c = ws.max_column
        if limit_to_col:
            try: max_c = min(max_c, _excel_col_to_idx(limit_to_col) + 1)
            except Exception: pass

        if header_row is None or data_start_row is None:
            scan = min(8, ws.max_row)
            best_r, best_nonempty = 1, -1
            for r in range(1, scan + 1):
                vals = [c.value for c in ws[r][0:max_c]]
                nonempty = sum(1 for x in vals if x not in (None, ""))
                if nonempty > best_nonempty:
                    best_nonempty = nonempty
                    best_r = r
            header_row = best_r
            data_start_row = best_r + 1

        raw_headers = [_format_cell(c) for c in ws[header_row][0:max_c]]
        raw_headers = [_norm_header_label(h) for h in raw_headers]
        headers = dedup(raw_headers)

        out_rows = []
        blanks_in_a_row = 0
        for r in range(int(data_start_row), ws.max_row + 1):
            cells = ws[r][0:max_c]
            row = [_format_cell(c) for c in cells]
            if all(v == "" for v in row):
                blanks_in_a_row += 1
                if blanks_in_a_row >= 3: break
                continue
            blanks_in_a_row = 0
            out_rows.append(row)

        df = pd.DataFrame(out_rows, columns=headers)
        df = df.dropna(how="all")
        df = df.replace("", np.nan).dropna(how="all").fillna("")
        df = df.loc[:, ~(df.astype(str).eq("").all())]
        return df
    finally:
        wb.close()

# ------------------------ site IDs (DK/FD) ------------------------

def _col_idx(c: Union[str, int]) -> int:
    # 0-based index from either "A"/"B"/... or int
    if isinstance(c, int): return max(0, c)
    s = re.sub(r"[^A-Za-z]", "", str(c)).upper()
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return max(0, n - 1)

def _salary_read_sheet(xlsm_path: Path, sheet: str, name_col, id_col,
                       team_col=None, pos_col=None, row_hard_cap: Optional[int]=None) -> List[Dict[str, str]]:
    """
    Fast reader for salary sheets. Stops after a long run of blank rows.
    """
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            return []

        ws = wb[sheet]
        name_i = _col_idx(name_col)
        id_i   = _col_idx(id_col)
        team_i = _col_idx(team_col) if team_col is not None else None
        pos_i  = _col_idx(pos_col)  if pos_col  is not None else None

        BLANK_BREAK = 200
        out, seen_ids = [], set()
        blank_run, seen_any = 0, False

        for r in range(1, ws.max_row + 1):
            name_v = ws.cell(r, name_i + 1).value
            id_v   = ws.cell(r, id_i + 1).value
            team_v = ws.cell(r, team_i + 1).value if team_i is not None else None
            pos_v  = ws.cell(r, pos_i + 1).value  if pos_i  is not None else None

            name = "" if name_v in (None, "Name") else str(name_v).strip()
            pid  = "" if id_v   in (None, "")      else str(id_v).strip()

            if name == "" and pid == "":
                if seen_any:
                    blank_run += 1
                    if blank_run >= BLANK_BREAK:
                        break
                continue

            blank_run = 0
            if not name or not pid:
                continue
            seen_any = True

            # Some files have "119110-85" etc. Keep digits and hyphen.
            pid_clean = (re.search(r"([\d\-]+)", pid) or [pid])[0]

            row = {"name": name, "id": pid_clean}
            if team_i is not None:
                row["team"] = str(team_v or "").strip().upper()
            if pos_i is not None:
                row["pos"]  = str(pos_v or "").strip().upper()
            if row["id"] in seen_ids:
                continue
            seen_ids.add(row["id"])
            out.append(row)

            if row_hard_cap and len(out) >= int(row_hard_cap):
                break

        return out
    finally:
        wb.close()

def run_site_ids(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any]) -> None:
    scfg = cfg.get("site_ids")
    if not scfg:
        print("⚠️  site_ids config missing — skipping.")
        return

    out_rel = (scfg.get("out_rel") or "").lstrip(r"\\/")
    if not out_rel:
        print("⚠️  site_ids.out_rel missing — skipping."); return

    dk_rows = _salary_read_sheet(
        xlsm_path,
        scfg.get("dk_sheet", "DK Salaries"),
        scfg.get("dk_name_col", "C"),
        scfg.get("dk_id_col",   "D"),
        scfg.get("dk_team_col"),
        scfg.get("dk_pos_col"),
        scfg.get("row_hard_cap"),
    )
    print(f"   DK site ids: {len(dk_rows)}")

    fd_rows = _salary_read_sheet(
        xlsm_path,
        scfg.get("fd_sheet", "FD Salaries"),
        scfg.get("fd_name_col", "D"),
        scfg.get("fd_id_col",   "A"),
        scfg.get("fd_team_col", "J"),   # your current layout
        scfg.get("fd_pos_col",  "B"),
        scfg.get("row_hard_cap"),
    )
    print(f"   FD site ids: {len(fd_rows)}")

    out = {"dk": dk_rows, "fd": fd_rows}
    out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
    ensure_parent(out_path)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✔️  JSON → {out_path}  (dk={len(dk_rows)}, fd={len(fd_rows)})")

# ------------------------ name crosswalk --------------------------

_SUFFIXES = {"jr","sr","ii","iii","iv","v"}

def _strip_accents(s):
    import unicodedata as u
    return u.normalize("NFKD", s).encode("ascii","ignore").decode("ascii")

def _norm_name(s: str) -> str:
    s = _strip_accents(str(s or "")).lower()
    s = re.sub(r"[^\w\s-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _base_key(s: str) -> str:
    parts = _norm_name(s).split()
    if parts and parts[-1].strip(".") in _SUFFIXES: parts = parts[:-1]
    return " ".join(parts)

def _fi_last(s: str) -> str:
    parts = _base_key(s).split()
    return (parts[0][:1] + " " + parts[-1]).strip() if parts else ""

def _last(s: str) -> str:
    parts = _base_key(s).split()
    return parts[-1] if parts else ""

def _build_index(rows: List[Dict[str, str]], name_key="name"):
    idx = {"by_norm": {}, "by_base": {}, "by_filast": {}, "rows": rows}
    for r in rows:
        nm = r.get(name_key, "")
        for key, bucket in ((_norm_name(nm),"by_norm"), (_base_key(nm),"by_base"), (_fi_last(nm),"by_filast")):
            if key:
                idx[bucket].setdefault(key, []).append(r)
    return idx

def _gate(cands: List[Dict[str,str]], team: str, pos: str) -> Optional[Dict[str,str]]:
    out = cands
    if team: out = [r for r in out if str(r.get("team","")).upper() == team]
    if pos:
        out2 = [r for r in out if str(r.get("pos","")).upper() == pos]
        if out2: out = out2
    return out[0] if len(out) == 1 else None

def _fuzzy_close(a: str, b: str, min_ratio=0.94) -> bool:
    import difflib
    return difflib.SequenceMatcher(None, _base_key(a), _base_key(b)).ratio() >= min_ratio

def run_name_xwalk(xlsm_path: Path, project_root: Path, cfg: Dict[str, Any]) -> None:
    nx = cfg.get("name_xwalk")
    if not nx:
        print("⚠️  name_xwalk config missing — skipping.")
        return

    out_rel = (nx.get("out_rel") or "").lstrip(r"\\/")
    if not out_rel:
        print("⚠️  name_xwalk.out_rel missing — skipping."); return

    sheet = nx.get("projections_sheet","Projections")
    df = read_literal_table(
        xlsm_path, sheet,
        nx.get("header_row", 1),
        nx.get("data_start_row", 2)
    )
    player_f = _resolve_col(df, nx.get("player_field","Player")) or "Player"
    team_f   = _resolve_col(df, nx.get("team_field","Team"))   or "Team"
    pos_f    = _resolve_col(df, nx.get("pos_field","Pos"))     or "Pos"

    scfg = cfg.get("site_ids") or {}
    dk_rows = _salary_read_sheet(
        xlsm_path, scfg.get("dk_sheet","DK Salaries"),
        scfg.get("dk_name_col","C"), scfg.get("dk_id_col","D"),
        scfg.get("dk_team_col"), scfg.get("dk_pos_col"),
        scfg.get("row_hard_cap"),
    )
    fd_rows = _salary_read_sheet(
        xlsm_path, scfg.get("fd_sheet","FD Salaries"),
        scfg.get("fd_name_col","D"), scfg.get("fd_id_col","A"),
        scfg.get("fd_team_col","J"), scfg.get("fd_pos_col","B"),
        scfg.get("row_hard_cap"),
    )

    dk_idx = _build_index(dk_rows)
    fd_idx = _build_index(fd_rows)

    out = []
    for _, r in df.iterrows():
        pname = str(r.get(player_f,"")).strip()
        if pname == "": continue
        team  = str(r.get(team_f,"")).upper()
        pos   = str(r.get(pos_f,"")).upper().split("/")[0]

        dk_hit = _gate(dk_idx["by_norm"].get(_norm_name(pname), []) or [], team, pos) \
              or _gate(dk_idx["by_base"].get(_base_key(pname), []) or [], team, pos) \
              or _gate(dk_idx["by_filast"].get(_fi_last(pname), []) or [], team, pos)

        fd_hit = _gate(fd_idx["by_norm"].get(_norm_name(pname), []) or [], team, pos) \
              or _gate(fd_idx["by_base"].get(_base_key(pname), []) or [], team, pos) \
              or _gate(fd_idx["by_filast"].get(_fi_last(pname), []) or [], team, pos)

        if not dk_hit:
            for rr in dk_idx["rows"]:
                if _last(rr.get("name","")) == _last(pname) and _fuzzy_close(pname, rr.get("name","")):
                    if (not team or rr.get("team","").upper()==team) and (not pos or rr.get("pos","").upper()==pos):
                        dk_hit = rr; break
        if not fd_hit:
            for rr in fd_idx["rows"]:
                if _last(rr.get("name","")) == _last(pname) and _fuzzy_close(pname, rr.get("name","")):
                    if (not team or rr.get("team","").upper()==team) and (not pos or rr.get("pos","").upper()==pos):
                        fd_hit = rr; break

        out.append({
            "proj": pname, "team": team, "pos": pos,
            "dk_name": (dk_hit or {}).get("name",""), "dk_id": (dk_hit or {}).get("id",""),
            "fd_name": (fd_hit or {}).get("name",""), "fd_id": (fd_hit or {}).get("id",""),
        })

    out_path = (project_root / "public" / Path(out_rel)).with_suffix(".json")
    ensure_parent(out_path)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✔️  JSON → {out_path}  (xwalk rows: {len(out)})")

# ----------------------------- main --------------------------------

def _choose_project_root(arg_proj: Optional[str]) -> Path:
    if arg_proj:
        p = Path(arg_proj).resolve()
        if (p / "public").exists():
            return p
        print(f"⚠️  --project '{p}' has no /public. Falling back to script root.", file=sys.stderr)
    if (ROOT / "public").exists():
        return ROOT
    print("ERROR: could not find a project root with /public.", file=sys.stderr)
    sys.exit(1)

def main() -> None:
    ap = argparse.ArgumentParser(description="Build weekly site_ids and name_xwalk (always rebuild).")
    ap.add_argument("--xlsm",    default=DEFAULT_XLSM,   help="Path to source workbook (.xls/.xlsx/.xlsm)")
    ap.add_argument("--project", default=DEFAULT_PROJ,   help="Path to project root (contains /public)")
    ap.add_argument("--config",  default=DEFAULT_CONFIG, help="Path to exporter config JSON")
    ap.add_argument("--only-site-ids", action="store_true", help="Only generate site_ids.json")
    ap.add_argument("--only-xwalk",    action="store_true", help="Only generate name_xwalk.json")
    args = ap.parse_args()

    xlsm_path    = Path(args.xlsm).resolve()
    project_root = _choose_project_root(args.project)
    config_path  = Path(args.config).resolve()

    if not xlsm_path.exists():
        print(f"ERROR: workbook not found: {xlsm_path}", file=sys.stderr); sys.exit(1)
    if not config_path.exists():
        print(f"ERROR: config not found: {config_path}", file=sys.stderr); sys.exit(1)

    staged_xlsm, temp_dir = _stage_copy_for_read(xlsm_path)
    try:
        cfg = json.loads(config_path.read_text(encoding="utf-8-sig"))

        if not args.only_xwalk:
            print("\n=== SITE IDS ===")
            run_site_ids(staged_xlsm, project_root, cfg)

        if not args.only_site_ids:
            print("\n=== NAME XWALK ===")
            run_name_xwalk(staged_xlsm, project_root, cfg)

        print("\nDone.")
    finally:
        try: shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception: pass

if __name__ == "__main__":
    main()
